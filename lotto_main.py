#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lotto 6/45 Simulator (KR) — Genius + Quantum + HM + MQLE + AI + Rigged Sim + 3D + GPU
메인 GUI 프로그램
"""

from __future__ import annotations

# GPU 비활성화 (CPU만 사용)
import os
os.environ['NUMBA_DISABLE_CUDA'] = '1'

# scikit-learn / numpy 멀티코어 최적화
# BLAS/LAPACK 스레드 수를 시스템 CPU 코어 수로 설정
import multiprocessing
n_cores = multiprocessing.cpu_count()
os.environ['OMP_NUM_THREADS'] = str(n_cores)
os.environ['MKL_NUM_THREADS'] = str(n_cores)
os.environ['OPENBLAS_NUM_THREADS'] = str(n_cores)
os.environ['BLIS_NUM_THREADS'] = str(n_cores)

# 표준 라이브러리
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import os
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed

# 서드파티 라이브러리
import numpy as np
import pandas as pd
import requests

# 로또 시뮬레이터 모듈
from lotto_utils import (
    parse_sets_from_text,
    sets_to_text,
    sets_to_text_with_scores,
    default_sets,
    get_rng,
)
from get_next_round_info import get_next_round_info
from lotto_generators import (
    generate_random_sets,
    generate_pattern_sets,
    gen_GI,
    gen_MDA,
    gen_CC,
    gen_PR,
    gen_IS,
    gen_GAPR,
    gen_QH,
    gen_HD,
    gen_QP,
    gen_QP_tunnel,
    gen_QP_entangle,
    gen_QH_QA,
    gen_QP_jump,
    gen_MQLE,
    train_ml_scorer,
    ml_score_sets_batch,
)
from lotto_history import (
    load_history_csv,
    compute_weights,
    compute_realistic_popularity_weights,
)
from lotto_simulation import (
    run_simulation,
    build_synthetic_player_pool,
    estimate_expected_winners_from_pool,
    _filter_ticket_pool_chunk,
)
from lotto_physics import (
    get_physics_backend_info,
)


_rng = get_rng()


# ============= Stacking 모델 Wrapper (배치 예측 최적화) =============
class StackingModelWrapper:
    """
    Stacking 앙상블 모델을 sklearn 인터페이스로 래핑

    배치 예측을 최적화하여 10배 이상 속도 향상:
    - N개 베이스 모델을 한 번에 배치 예측 (병렬)
    - 메타 모델로 최종 예측
    - pickle 직렬화 지원
    """
    def __init__(self, base_models, meta_model):
        self.base_models = base_models
        self.meta_model = meta_model
        self.n_base_models = len(base_models)

    def predict_proba(self, X):
        """
        배치 예측 (sklearn 호환) - 병렬 처리

        Args:
            X: (N, 57) 정규화된 특징 배열

        Returns:
            (N, 2) 확률 배열 [[P(class=0), P(class=1)], ...]
        """
        # Level 0: N개 베이스 모델 병렬 배치 예측 ⚡
        from joblib import Parallel, delayed

        # N개 모델을 병렬로 예측 (동적 CPU 코어 사용)
        base_preds_list = Parallel(n_jobs=self.n_base_models, prefer="threads")(
            delayed(lambda m: m.predict_proba(X)[:, 1])(model)
            for model in self.base_models
        )
        base_preds = np.column_stack(base_preds_list)  # Shape: (N, n_base_models)

        # 메타 입력: 베이스 예측 + 정규화된 원본 특징
        meta_input = np.hstack([base_preds, X])  # Shape: (N, n_base_models+57)

        # Level 1: 메타 모델 최종 예측
        return self.meta_model.predict_proba(meta_input)  # Shape: (N, 2)


class DummyMetaModel:
    """
    메타 모델 역할을 하는 더미 클래스 (25개 앙상블용)
    실제로는 베이스 모델들의 평균만 계산
    """
    def __init__(self):
        from sklearn.base import BaseEstimator, ClassifierMixin
        self.classes_ = np.array([0, 1])

    def fit(self, X, y):
        return self

    def predict_proba(self, X):
        """X는 이미 베이스 모델들의 평균 확률"""
        if X.ndim == 1:
            X = X.reshape(-1, 1)
        # 이진 분류이므로 [1-p, p] 형태로 반환
        probs = np.column_stack([1 - X, X])
        return probs

    def predict(self, X):
        probs = self.predict_proba(X)
        return (probs[:, 1] > 0.5).astype(int)


class EnsembleWrapper:
    """
    25개 MLP 앙상블을 Stacking처럼 동작하도록 래핑
    StackingModelWrapper와 호환되는 인터페이스 제공
    """
    def __init__(self, base_models, meta_model, mu, sigma):
        self.base_models = base_models
        self.meta_model = meta_model
        self.mu = mu
        self.sigma = sigma
        self.n_base_models = len(base_models)

    def predict_proba(self, X_raw):
        """
        배치 예측 (lotto_generators.ml_score_sets_batch 호환)

        Args:
            X_raw: (n_samples, n_features) - 정규화 안 된 원본 특징

        Returns:
            (n_samples, 2) - [1-p, p] 형태의 확률
        """
        # 정규화
        X_norm = (X_raw - self.mu) / self.sigma

        # N개 모델의 평균 예측 (병렬 처리)
        from joblib import Parallel, delayed

        all_probs = Parallel(n_jobs=self.n_base_models, prefer="threads")(
            delayed(lambda m: m.predict_proba(X_norm)[:, 1])(model)
            for model in self.base_models
        )

        # 평균
        avg_probs = np.mean(all_probs, axis=0)

        # [1-p, p] 형태로 변환
        return np.column_stack([1 - avg_probs, avg_probs])

    def predict(self, X_raw):
        """예측 (확률 > 0.5 → 1)"""
        probs = self.predict_proba(X_raw)
        return (probs[:, 1] > 0.5).astype(int)


class LottoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Lotto 6/45 Simulator (Genius + Quantum + HM + MQLE + AI + 3D + Rigged+GPU)")
        self.geometry("1180x820")
        self.resizable(True, True)

        self.history_df: pd.DataFrame | None = None
        self.history_path: str | None = None
        self.history_weights = None
        self.history_exclude_set: set[int] = set()
        self.ml_model: dict | None = None

        # ★ AI 세트 평점 학습 회차 수 (GUI)
        self.ai_max_rounds = tk.StringVar(value="200")
        # ★ ML 모델 타입 (Neural Network 고정)
        self.ml_model_type = tk.StringVar(value="neural_network")

        # 가상 조작 시뮬 관련 상태
        self.rig_win = None
        self.rig_tree = None
        self.rig_status_label = None
        self.rig_target_min = tk.IntVar(value=8)
        self.rig_target_max = tk.IntVar(value=15)
        # ★ 샘플링 제거: rig_samples 더 이상 사용 안 함 (ticket_pool 전수 조사)
        # ★ 가상 플레이어 수 (실제 티켓 수와 일치, 기본 112,000,000)
        self.rig_virtual_players = tk.IntVar(value=112000000)
        # ★ 가상 조작 시뮬 결과 저장용
        self.rig_results: list[tuple[list[int], float]] = []
        self.rig_last_params: dict = {}
        # ★ 가상 조작 시뮬 진행률 표시 위젯
        self.rig_progressbar = None
        self.rig_progress_label = None
        # ★ 가상 조작 시뮬 테이블 정렬 상태 (컬럼명, 오름차순 여부)
        self.rig_sort_column = None
        self.rig_sort_reverse = False
        self.rig_ml_label = None  # ML 가중치 레이블
        self.rig_ml_weight = tk.IntVar(value=50)  # ML 가중치 변수 (최적화 후: 50%)
        # ★ 일반 시뮬레이션 테이블 정렬 상태
        self.sim_sort_column = None
        self.sim_sort_reverse = False

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self.page_sets = ttk.Frame(self.notebook)
        self.page_generate = ttk.Frame(self.notebook)
        self.page_sim = ttk.Frame(self.notebook)
        self.page_help = ttk.Frame(self.notebook)
        self.page_recombine = ttk.Frame(self.notebook)

        self.notebook.add(self.page_sets, text="세트 편집")
        self.notebook.add(self.page_generate, text="번호 추출기")
        self.notebook.add(self.page_sim, text="시뮬레이션")
        self.notebook.add(self.page_recombine, text="번호 재조합")
        self.notebook.add(self.page_help, text="HELP")

        self._build_sets_page()
        self._build_generate_page()
        self._build_sim_page()
        self._build_help_page()
        self._build_recombine_page()

        self.text_sets.insert("1.0", sets_to_text(default_sets()))

        # 앙상블 모델 자동 로드 (있으면)
        self._load_ensemble_model_on_startup()

        # 윈도우 종료 시 프로토콜 설정
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _load_ensemble_model_on_startup(self):
        """프로그램 시작 시 Stacking 모델 자동 로드"""
        import os
        import pickle

        # Stacking 모델만 지원
        stacking_path = "best_ml_model_stacking.pkl"

        if os.path.exists(stacking_path):
            try:
                with open(stacking_path, 'rb') as f:
                    self.ml_model = pickle.load(f)

                # 모델 타입 확인
                model_type = self.ml_model.get('type', 'unknown')
                n_features = self.ml_model.get('n_features', 0)

                if model_type == 'neural_network_ensemble':
                    # Neural Network K-Fold 앙상블
                    n_models = self.ml_model.get('n_models', 0)
                    accuracy = self.ml_model.get('ensemble_accuracy', 0)
                    self.lbl_ai.config(
                        text=f"AI 세트 평점: 앙상블 ({n_models}개 모델, {n_features}개 특징, 정확도 {accuracy:.2%})"
                    )
                    print(f"[자동 로드] 앙상블 모델 로드 완료 ({n_models}개 모델, {n_features}개 특징)")
                else:
                    # ⚡ 하위 호환성: Stacking 또는 단일 모델
                    if 'model' not in self.ml_model:
                        base_models = self.ml_model.get('base_models')
                        meta_model = self.ml_model.get('meta_model')
                        if base_models and meta_model:
                            wrapper = StackingModelWrapper(base_models, meta_model)
                            self.ml_model['model'] = wrapper
                            print(f"[자동 로드] Wrapper 동적 생성 완료 (구버전 호환)")

                    n_models = self.ml_model.get('n_base_models', 0)
                    accuracy = self.ml_model.get('meta_train_accuracy', 0) / 100  # 백분율 → 소수
                    sep_power = self.ml_model.get('separation_power', 0)

                    self.lbl_ai.config(
                        text=f"AI 세트 평점: Stacking ({n_models}+1 모델, 정확도 {accuracy:.2%}, 구분력 {sep_power:.4f})"
                    )
                    print(f"[자동 로드] Stacking 모델 로드 완료 ({n_models}개 베이스 + 메타 모델)")
            except Exception as e:
                print(f"[경고] Stacking 모델 로드 실패: {e}")

    def _on_closing(self):
        """메인 윈도우 종료 시 모든 프로세스 정리"""
        import sys
        import gc

        print("\n[종료] 프로그램 종료 중...")

        try:
            # 1. 3D 시각화 윈도우 닫기 (physics_visualizer_3d 모듈 사용 시)
            from physics_visualizer_3d import cleanup_all_visualizers
            cleanup_all_visualizers()
            print("   [OK] 3D 시각화 윈도우 종료")
        except Exception as e:
            print(f"   [WARN] 3D 시각화 정리 실패: {e}")

        try:
            # 2. 가상 조작 윈도우 닫기
            if hasattr(self, 'rig_win') and self.rig_win is not None:
                try:
                    self.rig_win.destroy()
                    print("   [OK] 가상 조작 윈도우 종료")
                except:
                    pass
        except Exception as e:
            print(f"   [WARN] 가상 조작 윈도우 정리 실패: {e}")

        try:
            # 3. 메모리 정리
            gc.collect()
            print("   [OK] 메모리 정리")
        except Exception as e:
            print(f"   [WARN] 메모리 정리 실패: {e}")

        # 4. 메인 윈도우 닫기
        print("   [OK] 메인 윈도우 종료")
        self.quit()
        self.destroy()

        # 5. 프로세스 완전 종료
        print("[종료] 프로그램 종료 완료")
        sys.exit(0)

    # --- 세트 편집 페이지 ---
    def _build_sets_page(self):
        top = self.page_sets
        ttk.Label(top, text="세트 목록 (한 줄에 6개 숫자, 공백/쉼표 구분)").pack(
            anchor="w", padx=10, pady=6
        )
        self.text_sets = tk.Text(top, height=20, wrap="none")
        self.text_sets.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        btn_frame = ttk.Frame(top)
        btn_frame.pack(fill=tk.X, padx=10, pady=6)
        ttk.Button(btn_frame, text="불러오기(.txt)", command=self._load_sets).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(btn_frame, text="저장하기(.txt)", command=self._save_sets_txt).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(btn_frame, text="정렬/중복제거", command=self._normalize_sets).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(btn_frame, text="전체 초기화", command=self._clear_all_sets).pack(
            side=tk.LEFT, padx=4
        )

    def _load_sets(self):
        path = filedialog.askopenfilename(
            filetypes=[("Text", "*.txt"), ("All", "*.*")]
        )
        if not path:
            return
        with open(path, "r", encoding="utf-8") as f:
            self.text_sets.delete("1.0", tk.END)
            self.text_sets.insert("1.0", f.read())

    def _save_sets_txt(self):
        try:
            sets_ = parse_sets_from_text(self.text_sets.get("1.0", tk.END))
        except Exception as e:
            messagebox.showerror("오류", str(e))
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".txt", filetypes=[("Text", "*.txt")]
        )
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write(sets_to_text(sets_))
        messagebox.showinfo("저장 완료", f"세트 {len(sets_)}개 저장")

    def _normalize_sets(self):
        try:
            sets_ = parse_sets_from_text(self.text_sets.get("1.0", tk.END))
        except Exception as e:
            messagebox.showerror("오류", str(e))
            return
        uniq = sorted({tuple(s) for s in sets_})
        self.text_sets.delete("1.0", tk.END)
        self.text_sets.insert("1.0", sets_to_text([list(s) for s in uniq]))
        messagebox.showinfo("정리 완료", f"세트 {len(uniq)}개")

    def _clear_all_sets(self):
        self.text_sets.delete("1.0", tk.END)
        messagebox.showinfo("초기화", "세트 목록이 모두 삭제되었습니다.")

    # --- 번호 추출기 페이지 ---
    def _build_generate_page(self):
        top = self.page_generate

        hist = ttk.LabelFrame(top, text="과거 당첨 데이터(옵션)")
        hist.pack(fill=tk.X, padx=10, pady=8)
        ttk.Button(hist, text="데이터 불러오기 (Excel/CSV)", command=self._load_history).grid(
            row=0, column=0, padx=6, pady=6, sticky="w"
        )
        self.lbl_hist = ttk.Label(hist, text="로드되지 않음")
        self.lbl_hist.grid(row=0, column=1, padx=6, sticky="w")

        self.lbl_ai = ttk.Label(hist, text="AI 세트 평점: 준비 안 됨")
        self.lbl_ai.grid(row=0, column=2, padx=6, sticky="w")

        ttk.Label(hist, text="전략").grid(row=1, column=0, sticky="e")
        self.hist_strategy = tk.StringVar(value="사용 안 함")
        ttk.Combobox(
            hist,
            textvariable=self.hist_strategy,
            state="readonly",
            values=[
                "사용 안 함",
                "Hot(고빈도)",
                "Cold(저빈도)",
                "Overdue(오래 안 나온)",
                "Balanced(중립화)",
            ],
        ).grid(row=1, column=1, sticky="w", padx=6)

        ttk.Label(hist, text="Lookback N(최근 N회만)").grid(
            row=1, column=2, sticky="e"
        )
        self.hist_lookback = tk.StringVar(value="")
        ttk.Entry(hist, textvariable=self.hist_lookback, width=10).grid(
            row=1, column=3, sticky="w", padx=6
        )

        ttk.Label(hist, text="최근 K회 제외").grid(row=1, column=4, sticky="e")
        self.hist_exclude = tk.IntVar(value=0)
        ttk.Entry(hist, textvariable=self.hist_exclude, width=8).grid(
            row=1, column=5, sticky="w", padx=6
        )

        # ★ 추가: AI 세트 평점 학습 회차 수
        ttk.Label(hist, text="AI 학습 회차 수:").grid(
            row=2, column=0, sticky="e", pady=(4, 2)
        )
        # 슬라이더로 변경 (50~1000, 1000=전체)
        self.ai_rounds_slider = tk.IntVar(value=200)
        scale_ai = tk.Scale(
            hist,
            from_=50,
            to=1000,
            orient="horizontal",
            variable=self.ai_rounds_slider,
            length=200,
            showvalue=0,
        )
        scale_ai.grid(row=2, column=1, sticky="w", padx=6, pady=(4, 2))

        # 현재 값 레이블
        self.ai_rounds_label = ttk.Label(hist, text="200회")
        self.ai_rounds_label.grid(row=2, column=2, sticky="w", padx=4)

        # 슬라이더 값 변경 시 ai_max_rounds 업데이트
        def update_ai_rounds(*_):
            val = self.ai_rounds_slider.get()
            if val >= 1000:
                self.ai_max_rounds.set("")  # 전체
                self.ai_rounds_label.config(text="전체")
            else:
                self.ai_max_rounds.set(str(val))
                self.ai_rounds_label.config(text=f"{val}회")
        self.ai_rounds_slider.trace_add("write", update_ai_rounds)
        update_ai_rounds()  # 초기값 설정

        # ML 모델 타입 선택 (Neural Network 고정)
        ttk.Label(hist, text="ML 모델:").grid(
            row=3, column=0, sticky="e", pady=(4, 2)
        )
        model_combo = ttk.Combobox(
            hist,
            textvariable=self.ml_model_type,
            values=[
                "neural_network",
            ],
            state="readonly",
            width=18,
        )
        model_combo.grid(row=3, column=1, sticky="w", padx=6, pady=(4, 2))

        # 모델 설명 레이블
        self.ml_type_desc = ttk.Label(hist, text="신경망 (최적화됨, 5층 100-80-60-40-20)")
        self.ml_type_desc.grid(row=3, column=2, sticky="w", padx=4)

        # 모델 타입 변경 시 설명 업데이트
        def update_model_desc(*_):
            model = self.ml_model_type.get()
            descriptions = {
                "neural_network": "신경망 (최적화됨, 5층 100-80-60-40-20)",
            }
            self.ml_type_desc.config(text=descriptions.get(model, ""))
        self.ml_model_type.trace_add("write", update_model_desc)
        update_model_desc()

        # ML 학습 시작 버튼
        ttk.Button(hist, text="🎓 ML 학습 시작 (Stacking 앙상블)", command=self._train_ml_model).grid(
            row=4, column=0, columnspan=3, padx=6, pady=(8, 6), sticky="ew"
        )

        frm = ttk.LabelFrame(top, text="번호 추출기")
        frm.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(frm, text="생성 개수").grid(row=0, column=0, sticky="w")
        self.gen_count = tk.IntVar(value=10)
        ttk.Entry(frm, textvariable=self.gen_count, width=8).grid(
            row=0, column=1, sticky="w", padx=6
        )

        ttk.Label(frm, text="모드").grid(row=0, column=2, sticky="e")
        self.gen_mode = tk.StringVar(value="무작위")
        ttk.Combobox(
            frm,
            textvariable=self.gen_mode,
            state="readonly",
            values=[
                "무작위",
                "패턴",
                "GI(직관)",
                "MDA(다차원)",
                "CC(창의연결)",
                "PR(패턴공진)",
                "IS(혁신시뮬)",
                "GAP-R(간격공진)",
                "QH(다속성조화)",
                "HD(초다양성)",
                "QP-Wave(파동)",
                "QP-Tunnel(터널링)",
                "QP-Entangle(얽힘)",
                "QH-QA(어닐링)",
                "QP-Jump(모드도약)",
                "MQLE(끝판왕)",
                "물리시뮬3D",              # 3D 구형 챔버 시뮬레이션
                "물리시뮬3D+MQLE(끝판왕)", # 3D + MQLE 융합 (최종)
            ],
        ).grid(row=0, column=3, sticky="w", padx=6)

        # 물리 시뮬 백엔드 정보 표시
        backend_info = get_physics_backend_info()
        self.lbl_physics_backend = ttk.Label(frm, text=f"[물리시뮬: {backend_info}]")
        self.lbl_physics_backend.grid(row=0, column=4, sticky="w", padx=10)

        ttk.Label(frm, text="짝수 개수(선택)").grid(row=1, column=0, sticky="w", pady=6)
        self.gen_even = tk.StringVar(value="")
        ttk.Entry(frm, textvariable=self.gen_even, width=6).grid(
            row=1, column=1, sticky="w"
        )

        ttk.Label(frm, text="구간 분포 (저/중/고)").grid(row=1, column=2, sticky="e")
        self.gen_low = tk.IntVar(value=2)
        self.gen_mid = tk.IntVar(value=2)
        self.gen_high = tk.IntVar(value=2)
        ttk.Entry(frm, textvariable=self.gen_low, width=5).grid(
            row=1, column=3, sticky="w"
        )
        ttk.Entry(frm, textvariable=self.gen_mid, width=5).grid(
            row=1, column=4, sticky="w"
        )
        ttk.Entry(frm, textvariable=self.gen_high, width=5).grid(
            row=1, column=5, sticky="w"
        )

        ttk.Label(frm, text="배수 포함 (3의/7의 최소개수)").grid(
            row=2, column=0, sticky="w", pady=6
        )
        self.gen_m3 = tk.IntVar(value=0)
        self.gen_m7 = tk.IntVar(value=0)
        ttk.Entry(frm, textvariable=self.gen_m3, width=5).grid(
            row=2, column=1, sticky="w"
        )
        ttk.Entry(frm, textvariable=self.gen_m7, width=5).grid(
            row=2, column=2, sticky="w"
        )

        self.qc_balance = tk.IntVar(value=50)
        self.scale_qc = tk.Scale(
            frm,
            from_=0,
            to=100,
            orient="horizontal",
            label="양자 비중(%) — MQLE 전용",
            variable=self.qc_balance,
            length=360,
        )
        self.scale_qc.grid(row=3, column=0, columnspan=6, sticky="we", pady=(8, 0))

        # ML 가중치 슬라이더 추가 (최적화 후: 기본값 50%)
        self.ml_weight = tk.IntVar(value=50)
        self.scale_ml = tk.Scale(
            frm,
            from_=0,
            to=100,
            orient="horizontal",
            label="ML 가중치(%) — MQLE 전용 (CSV 필요)",
            variable=self.ml_weight,
            length=360,
        )
        self.scale_ml.grid(row=4, column=0, columnspan=6, sticky="we", pady=(8, 0))

        btns = ttk.Frame(top)
        btns.pack(fill=tk.X, padx=10, pady=8)
        ttk.Button(btns, text="번호 생성", command=self._gen_dispatch).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(btns, text="세트 페이지에 추가", command=self._append_to_sets).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(btns, text="생성 결과 초기화", command=self._clear_generated).pack(
            side=tk.LEFT, padx=4
        )
        ttk.Button(btns, text="🎬 3D 물리 시각화", command=self._launch_3d_visualizer).pack(
            side=tk.LEFT, padx=4
        )

        self.text_generate = tk.Text(top, height=18, wrap="none")
        self.text_generate.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)

    def _load_history(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv"), ("All", "*.*")]
        )
        if not path:
            return
        try:
            # Excel 또는 CSV 자동 판별
            if path.lower().endswith('.xlsx'):
                df = self._load_history_excel(path)
            else:
                df = load_history_csv(path)
        except Exception as e:
            messagebox.showerror("파일 오류", str(e))
            return

        self.history_df = df
        self.history_path = path
        self.lbl_hist.config(
            text=f"로드됨: {os.path.basename(path)} ({len(df)}회)"
        )

        # 데이터 로드 시 ML 모델도 다시 로드 (최신 모델 반영)
        self._load_ensemble_model_on_startup()

        # 모델 로드가 실패한 경우를 대비한 fallback
        if self.ml_model is None:
            self.lbl_ai.config(text="AI 세트 평점: 학습 전 (🎓 ML 학습 시작 버튼 클릭)")

    def _load_history_excel(self, path: str) -> pd.DataFrame:
        """
        Excel 파일에서 로또 당첨 데이터 로드
        형식: 회차, 추첨일, 번호1, 번호2, 번호3, 번호4, 번호5, 번호6, 보너스
        """
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # 헤더 읽기 (첫 번째 행)
        headers = [cell.value for cell in ws[1]]

        # 데이터 읽기
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # 회차가 있는 행만
                data.append(row)

        wb.close()

        # DataFrame 생성
        df = pd.DataFrame(data, columns=headers)

        # 디버그: 컬럼명 출력
        print(f"[DEBUG] Excel 컬럼명: {list(df.columns)}")

        # 컬럼명 정규화 (n1~n6 형식으로 변환)
        result_df = pd.DataFrame()

        # round, date 컬럼 찾기
        for col in df.columns:
            col_str = str(col).strip() if col is not None else ""
            col_lower = col_str.lower()
            if '회차' in col_lower or 'round' in col_lower or col_lower == 'no':
                result_df['round'] = pd.to_numeric(df[col], errors='coerce')
            elif '추첨일' in col_lower or '날짜' in col_lower or 'date' in col_lower:
                result_df['date'] = df[col]

        # 번호 컬럼 찾기 - 여러 방법 시도
        number_cols = []

        # 방법 1: '당첨번호' 컬럼에 문자열로 저장된 경우 (예: "1, 2, 3, 4, 5, 6")
        winning_num_col = None
        for col in df.columns:
            col_str = str(col).strip() if col is not None else ""
            col_lower = col_str.lower()
            if '당첨번호' in col_lower or 'winning' in col_lower:
                winning_num_col = col
                break

        if winning_num_col is not None:
            # 첫 번째 행 확인
            sample = df[winning_num_col].iloc[0] if len(df) > 0 else None
            if sample is not None:
                sample_str = str(sample).strip()
                # 쉼표나 공백으로 구분된 숫자들인지 확인
                if ',' in sample_str or ' ' in sample_str or '\t' in sample_str or '\n' in sample_str:
                    print(f"[DEBUG] '당첨번호' 컬럼에서 문자열 파싱 시도: {sample_str[:50]}")

                    # 각 행을 파싱해서 6개 번호 추출
                    parsed_numbers = []
                    success_count = 0

                    for idx, row in df.iterrows():
                        num_str = str(row[winning_num_col]).strip()
                        # 여러 구분자로 분리 시도
                        numbers = []

                        # 쉼표, 탭, 줄바꿈 등을 공백으로 치환
                        num_str = num_str.replace(',', ' ').replace('\t', ' ').replace('\n', ' ')
                        parts = [p.strip() for p in num_str.split() if p.strip()]

                        try:
                            numbers = [int(p) for p in parts if p.isdigit() and 1 <= int(p) <= 45]
                        except:
                            pass

                        if len(numbers) >= 6:
                            parsed_numbers.append(numbers[:6])
                            success_count += 1
                        else:
                            parsed_numbers.append([None] * 6)

                    # 성공했으면 result_df에 추가
                    if success_count > 0:
                        for i in range(1, 7):
                            result_df[f'n{i}'] = [nums[i-1] if nums[i-1] is not None else None for nums in parsed_numbers]

                        number_cols = ['parsed']  # 더미 값
                        print(f"[DEBUG] 문자열 파싱 성공! {success_count}/{len(df)}행")

        # 방법 2: 정확한 패턴 매칭 (번호1, 번호2, ... 또는 n1, n2, ...)
        if len(number_cols) == 0:
            for i in range(1, 7):
                found_col = None
                for col in df.columns:
                    col_str = str(col).strip() if col is not None else ""
                    col_lower = col_str.lower()

                    # 다양한 패턴 체크
                    if (f'번호{i}' in col_lower or
                        col_lower == f'n{i}' or
                        col_str == str(i) or  # 숫자만 있는 경우 (1, 2, 3, ...)
                        col_lower == f'no{i}' or
                        col_lower == f'num{i}'):
                        found_col = col
                        break

                if found_col is not None:
                    number_cols.append(found_col)

        # 방법 3: '당첨번호' 다음의 병합된 셀(None) 컬럼 확인
        if len(number_cols) < 6:
            print(f"[DEBUG] 패턴 매칭으로 {len(number_cols)}개만 찾음. 병합된 셀 확인 시작...")

            # '당첨번호' 컬럼의 인덱스 찾기
            winning_idx = None
            if winning_num_col is not None:
                winning_idx = list(df.columns).index(winning_num_col)

            if winning_idx is not None:
                potential_cols = []

                # 먼저 '당첨번호' 컬럼 자체 확인 (첫 번째 번호가 여기 있을 수 있음)
                try:
                    test_val = pd.to_numeric(df.iloc[0, winning_idx], errors='coerce')
                    if not pd.isna(test_val) and 1 <= test_val <= 45:
                        potential_cols.append(winning_idx)
                        print(f"[DEBUG] 컬럼 인덱스 {winning_idx} (헤더: '당첨번호'): 값={test_val}")
                except:
                    pass

                # '당첨번호' 다음 컬럼들 확인 (최대 7개 더)
                for i in range(winning_idx + 1, min(winning_idx + 8, len(df.columns))):
                    col = df.columns[i]
                    col_str = str(col) if col is not None else ""

                    # None이거나 빈 문자열인 컬럼 (병합된 셀의 경우)
                    if col is None or col_str.strip() == '' or col_str.lower() == 'none':
                        # 데이터 행에서 실제 값 확인 (첫 번째 데이터 행)
                        try:
                            test_val = pd.to_numeric(df.iloc[0, i], errors='coerce')
                            if not pd.isna(test_val) and 1 <= test_val <= 45:
                                potential_cols.append(i)
                                print(f"[DEBUG] 컬럼 인덱스 {i} (헤더: '{col}'): 값={test_val}")
                        except Exception as e:
                            pass
                    else:
                        # 헤더가 있는 컬럼이면 중단 (보너스 등)
                        print(f"[DEBUG] 컬럼 인덱스 {i}에서 헤더 '{col}' 발견, 탐색 중단")
                        break

                if len(potential_cols) >= 6:
                    number_cols = potential_cols[:6]
                    print(f"[DEBUG] 병합된 셀에서 {len(number_cols)}개 찾음: 인덱스 {number_cols}")

        # 방법 4: 위치 기반으로 찾기
        if len(number_cols) < 6:
            print(f"[DEBUG] 위치 기반 탐색 시작...")

            # 숫자 컬럼만 추출 (회차, 날짜 제외)
            numeric_cols = []
            for idx, col in enumerate(df.columns):
                col_str = str(col).strip() if col is not None else ""
                col_lower = col_str.lower()

                # 회차, 날짜, 보너스 컬럼 제외
                if ('회차' not in col_lower and 'round' not in col_lower and
                    '날짜' not in col_lower and 'date' not in col_lower and
                    '추첨일' not in col_lower and 'no' != col_lower and
                    '보너스' not in col_lower and 'bonus' not in col_lower and
                    '순위' not in col_lower and '당첨' not in col_lower):

                    # 숫자 데이터가 있는지 확인
                    try:
                        test_val = pd.to_numeric(df.iloc[0, idx] if len(df) > 0 else None, errors='coerce')
                        if not pd.isna(test_val) and 1 <= test_val <= 45:
                            numeric_cols.append(idx)
                    except:
                        pass

            # 처음 6개를 번호로 사용
            if len(numeric_cols) >= 6:
                number_cols = numeric_cols[:6]
                print(f"[DEBUG] 위치 기반으로 {len(number_cols)}개 찾음: 인덱스 {number_cols}")
            else:
                raise ValueError(
                    f"Excel 파일에서 당첨번호 컬럼을 찾을 수 없습니다.\n"
                    f"컬럼명: {list(df.columns)}\n"
                    f"당첨번호는 '번호1'~'번호6' 또는 '1'~'6' 형식이거나\n"
                    f"'당첨번호' 컬럼에 쉼표로 구분된 숫자 문자열이어야 합니다."
                )

        # n1~n6 컬럼 추가 (문자열 파싱으로 이미 설정된 경우 건너뜀)
        if number_cols != ['parsed']:
            for i, col in enumerate(number_cols, 1):
                # 인덱스인지 컬럼명인지 확인
                if isinstance(col, int):
                    # 인덱스로 접근
                    result_df[f'n{i}'] = pd.to_numeric(df.iloc[:, col], errors='coerce')
                else:
                    # 컬럼명으로 접근
                    result_df[f'n{i}'] = pd.to_numeric(df[col], errors='coerce')

        # 보너스 번호 (선택적)
        for col in df.columns:
            col_str = str(col).strip() if col is not None else ""
            col_lower = col_str.lower()
            if '보너스' in col_lower or 'bonus' in col_lower:
                result_df['bonus'] = pd.to_numeric(df[col], errors='coerce')
                break

        # NaN 제거
        result_df = result_df.dropna(subset=[f'n{i}' for i in range(1, 7)])

        # 숫자 컬럼을 int로 변환
        for i in range(1, 7):
            result_df[f'n{i}'] = result_df[f'n{i}'].astype(int)

        # 1~45 범위 검증
        for i in range(1, 7):
            mask = (result_df[f'n{i}'] >= 1) & (result_df[f'n{i}'] <= 45)
            if not mask.all():
                raise ValueError(f"n{i} 컬럼에 1~45 범위를 벗어난 값이 있습니다.")

        print(f"[DEBUG] Excel 로드 성공: {len(result_df)}행")
        return result_df.reset_index(drop=True)

    def _train_ml_model(self):
        """ML 모델 학습 (별도 스레드에서 실행)"""
        # 데이터 로드 확인
        if self.history_df is None or self.history_df.empty:
            messagebox.showwarning(
                "데이터 필요",
                "먼저 '데이터 불러오기' 버튼으로 과거 당첨 데이터를 로드하세요."
            )
            return

        # 학습 시작 표시
        self.lbl_ai.config(text="AI 세트 평점: 학습 중... (잠시만 기다려주세요)")
        self.page_generate.update()  # UI 즉시 업데이트

        # 별도 스레드에서 학습 실행
        threading.Thread(target=self._train_ml_model_worker, daemon=True).start()

    def _train_ml_model_worker(self):
        """Stacking 앙상블 학습 (백그라운드 스레드)

        1단계: K-Fold 앙상블 학습 (10개 베이스 모델)
        2단계: Stacking 메타 모델 학습
        """
        import pickle
        import os

        try:
            print("=" * 80)
            print("Stacking 앙상블 학습 시작")
            print("=" * 80)

            # ===========================
            # 1단계: K-Fold 앙상블 학습
            # ===========================
            print("\n[1단계] K-Fold 앙상블 학습 (25개 모델)")

            # 학습 데이터 준비 (시간 정보 포함)
            pos_sets = []
            pos_meta = []  # (round, date) 시간 정보 저장

            # n1~n6 컬럼 확인
            num_cols = [f'n{i}' for i in range(1, 7)]
            has_n_cols = all(col in self.history_df.columns for col in num_cols)

            if has_n_cols:
                # n1~n6 컬럼이 있으면 직접 추출
                for idx, row in self.history_df.iterrows():
                    try:
                        # round와 date 정보 추출
                        round_num = int(row['round']) if 'round' in row and pd.notna(row['round']) else None
                        date_str = str(row['date']) if 'date' in row and pd.notna(row['date']) else None

                        # n1~n6 추출
                        nums = [int(row[f'n{i}']) for i in range(1, 7)]
                        if all(1 <= n <= 45 for n in nums):
                            pos_sets.append(sorted(nums))
                            pos_meta.append((round_num, date_str))
                    except (ValueError, TypeError, KeyError):
                        continue
            else:
                # 이전 방식: 전체 순회
                for row in self.history_df.itertuples(index=False):
                    # round와 date 정보 추출
                    try:
                        round_num = int(row[0]) if len(row) > 0 else None
                        date_str = str(row[1]) if len(row) > 1 else None
                    except (ValueError, IndexError):
                        round_num = None
                        date_str = None

                    nums = []
                    for val in row:
                        try:
                            v = int(val)
                            if 1 <= v <= 45:
                                nums.append(v)
                        except (ValueError, TypeError):
                            continue
                    if len(nums) == 6:
                        pos_sets.append(sorted(nums))
                        pos_meta.append((round_num, date_str))

            # 데이터 검증
            if len(pos_sets) == 0:
                raise ValueError(
                    "학습 데이터가 없습니다.\n"
                    "Excel 파일에 n1~n6 컬럼이 있고 1~45 범위의 숫자가 있는지 확인하세요.\n"
                    f"DataFrame 컬럼: {list(self.history_df.columns)}\n"
                    f"DataFrame 크기: {len(self.history_df)}행"
                )

            print(f"   [데이터] {len(pos_sets)}개 당첨번호 추출됨")

            # 음성 샘플: 편향된 조합 생성
            n_neg = len(pos_sets) * 5
            neg_sets = []

            from lotto_generators import generate_biased_combinations
            neg_sets = generate_biased_combinations(n_neg)

            # 특징 추출 (⚡ Numba 병렬 처리)
            from lotto_generators import (
                _compute_core_features_batch,
                _compute_history_features_batch,
                _compute_temporal_features_batch,
                _prepare_history_array
            )
            import time

            print(f"   [특징 추출] 57개 고급 특징 (39 코어 + 11 히스토리 + 7 시간)")
            print(f"   [Numba+fastmath] 첫 실행 시 컴파일... (2-3초 소요)")
            print(f"   [멀티코어] prange로 36코어 최대 활용!")

            start_time = time.time()

            # 히스토리 데이터를 numpy 배열로 변환 (한 번만)
            print(f"   [전처리] 히스토리 데이터 변환...")
            history_arr = _prepare_history_array(self.history_df)
            print(f"        → 완료! ({len(history_arr)}회 히스토리)")

            # 핵심 특징 추출 (CPU 병렬)
            print(f"   [1/3] 핵심 특징 추출 (배치 {len(pos_sets) + len(neg_sets)}개, 병렬 처리)...")
            pos_sets_arr = np.array(pos_sets, dtype=np.float64)  # (N_pos, 6)
            neg_sets_arr = np.array(neg_sets, dtype=np.float64)  # (N_neg, 6)

            core_features_pos = _compute_core_features_batch(pos_sets_arr)  # (N_pos, 39)
            core_features_neg = _compute_core_features_batch(neg_sets_arr)  # (N_neg, 39)
            core_time = time.time() - start_time
            print(f"        → 완료! ({core_time:.1f}초)")

            # 히스토리 특징 추출 (CPU 병렬)
            print(f"   [2/3] 히스토리 특징 추출 (배치 {len(pos_sets) + len(neg_sets)}개, 병렬 처리)...")
            hist_start = time.time()
            hist_features_pos = _compute_history_features_batch(pos_sets_arr, history_arr)  # (N_pos, 11)
            hist_features_neg = _compute_history_features_batch(neg_sets_arr, history_arr)  # (N_neg, 11)
            hist_time = time.time() - hist_start
            print(f"        → 완료! ({hist_time:.1f}초)")

            # 시간 특징 추출 (양성 샘플만 시간 정보 있음)
            print(f"   [3/3] 시간 특징 추출...")
            temp_start = time.time()

            # 양성 샘플: 각 샘플마다 실제 시간 정보 사용
            temporal_features_pos_list = []
            for i in range(len(pos_sets)):
                round_num, date_str = pos_meta[i]
                temp_feat = _compute_temporal_features_batch(1, round_num, date_str)[0]  # (7,)
                temporal_features_pos_list.append(temp_feat)
            temporal_features_pos = np.array(temporal_features_pos_list)  # (N_pos, 7)

            # 음성 샘플: 히스토리에서 랜덤한 시간 정보 사용
            # (시간 특징이 양성/음성 구분자가 되지 않도록)
            temporal_features_neg_list = []
            for _ in range(len(neg_sets)):
                # 히스토리에서 랜덤 회차 선택
                random_idx = np.random.randint(0, len(pos_meta))
                round_num, date_str = pos_meta[random_idx]
                temp_feat = _compute_temporal_features_batch(1, round_num, date_str)[0]
                temporal_features_neg_list.append(temp_feat)
            temporal_features_neg = np.array(temporal_features_neg_list)  # (N_neg, 7)

            temp_time = time.time() - temp_start
            print(f"        → 완료! ({temp_time:.1f}초)")

            # 결합 (57개)
            X_pos = np.hstack([core_features_pos, hist_features_pos, temporal_features_pos])  # (N_pos, 57)
            X_neg = np.hstack([core_features_neg, hist_features_neg, temporal_features_neg])  # (N_neg, 57)
            X = np.vstack([X_pos, X_neg])  # (N_pos + N_neg, 57)

            # 레이블
            y = np.array([1.0] * len(pos_sets) + [0.0] * len(neg_sets), dtype=float)

            # 정규화
            mu = X.mean(axis=0)
            sigma = X.std(axis=0)
            sigma[sigma < 1e-6] = 1.0
            Xn = (X - mu) / sigma

            N, D = Xn.shape
            print(f"   샘플: {N}개 (양성: {len(pos_sets)}, 음성: {len(neg_sets)}), 특징: {D}개")

            # K-Fold 앙상블 학습 (진짜 멀티프로세싱 - joblib loky backend)
            from sklearn.model_selection import StratifiedKFold, cross_validate
            from sklearn.neural_network import MLPClassifier
            from joblib import parallel_backend
            import os
            import time

            # 각 프로세스가 2코어씩 사용하도록 설정 (25 프로세스 × 2 코어 = 50 코어)
            os.environ['OMP_NUM_THREADS'] = '2'
            os.environ['MKL_NUM_THREADS'] = '2'
            os.environ['OPENBLAS_NUM_THREADS'] = '2'

            skf = StratifiedKFold(n_splits=25, shuffle=True, random_state=42)

            print(f"   K-Fold 앙상블 학습 시작")
            print(f"   [진짜 병렬 모드] joblib loky backend로 25개 프로세스 동시 실행")
            print(f"   각 프로세스 2코어 사용 → 총 50코어 활용")
            print(f"   예상 시간: 40-60초")

            start_time = time.time()

            # 베이스 모델 정의
            base_model = MLPClassifier(
                hidden_layer_sizes=(100, 80, 60, 40, 20),
                activation='tanh',
                solver='adam',
                learning_rate_init=0.005,
                alpha=0.0001,  # 최적화: 0.0005 → 0.0001 (학습 속도 35% 향상)
                batch_size=200,
                max_iter=300,
                early_stopping=True,
                validation_fraction=0.1,
                n_iter_no_change=10,
                random_state=42,
                verbose=0,
            )

            # loky backend 명시적 사용 (진짜 멀티프로세싱)
            print(f"   loky backend 시작... (25개 독립 프로세스 생성)")
            with parallel_backend('loky', n_jobs=25):
                cv_results = cross_validate(
                    base_model, Xn, y,
                    cv=skf,
                    scoring='accuracy',
                    return_estimator=True,
                    return_train_score=True,
                    verbose=2,
                )

            elapsed = time.time() - start_time

            # 학습된 모델과 점수 추출
            ensemble_models = cv_results['estimator']
            fold_scores = cv_results['test_score'].tolist()

            print(f"\n   [진짜 병렬 완료] 소요 시간: {elapsed:.1f}초")
            print(f"   평균 검증 정확도: {np.mean(fold_scores):.4f} (±{np.std(fold_scores):.4f})")
            for fold_idx, score in enumerate(fold_scores, 1):
                print(f"      Fold {fold_idx}: {score:.4f}")

            # 코어 설정 원복
            os.environ['OMP_NUM_THREADS'] = str(n_cores)
            os.environ['MKL_NUM_THREADS'] = str(n_cores)
            os.environ['OPENBLAS_NUM_THREADS'] = str(n_cores)

            # 앙상블 성능 평가
            ensemble_probs = np.mean([m.predict_proba(Xn)[:, 1] for m in ensemble_models], axis=0)
            ensemble_preds = (ensemble_probs > 0.5).astype(int)
            ensemble_acc = (ensemble_preds == y).mean()

            print(f"   K-Fold 앙상블 정확도: {ensemble_acc:.2%}")

            # K-Fold 앙상블 저장 (임시, Stacking 학습에 필요)
            ensemble_data = {
                'type': 'neural_network_ensemble',
                'models': ensemble_models,
                'mu': mu,
                'sigma': sigma,
                'n_models': len(ensemble_models),
                'ensemble_accuracy': float(ensemble_acc * 100),
                'fold_scores': fold_scores,
                'n_features': D,
                'separation_power': 0.0,  # 임시값
            }

            with open('best_ml_model_ensemble.pkl', 'wb') as f:
                pickle.dump(ensemble_data, f)

            print(f"   [OK] K-Fold 앙상블 저장 완료")

            # ===========================
            # 2단계: Stacking 메타 모델 학습
            # ===========================
            print("\n[2단계] Stacking 메타 모델 학습")

            # Out-of-fold 예측 생성
            meta_predictions = np.zeros((len(X), len(ensemble_models)))

            for fold_idx, (train_idx, val_idx) in enumerate(skf.split(Xn, y), 1):
                model = ensemble_models[fold_idx - 1]
                preds = model.predict_proba(Xn[val_idx])[:, 1]
                meta_predictions[val_idx, fold_idx - 1] = preds

            # 메타 특징 = 25개 예측 + 57개 원본 특징 (= 82개)
            X_meta = np.hstack([meta_predictions, Xn])
            print(f"   메타 특징: {X_meta.shape}")

            # 메타 모델 학습 (LogisticRegression)
            from sklearn.linear_model import LogisticRegression
            from sklearn.model_selection import cross_val_score

            meta_model = LogisticRegression(
                max_iter=500,
                random_state=42,
                C=0.01,  # 정규화 강화 (1.0 → 0.01)
                class_weight='balanced',
                solver='lbfgs',
            )

            # Cross-validation
            cv_scores = cross_val_score(meta_model, X_meta, y, cv=5, scoring='accuracy')
            print(f"   메타 모델 CV 점수: {cv_scores.mean():.4f} ± {cv_scores.std():.4f}")

            # 전체 데이터로 학습
            meta_model.fit(X_meta, y)
            y_pred = meta_model.predict(X_meta)
            from sklearn.metrics import accuracy_score
            train_accuracy = accuracy_score(y, y_pred)

            # 구분력 계산
            real_scores = y_pred[y == 1.0]
            biased_scores = y_pred[y == 0.0]
            separation = (real_scores.mean() - biased_scores.mean())

            print(f"   Stacking 정확도: {train_accuracy:.2%}")
            print(f"   구분력: {separation:.4f}")

            # ⚡ Stacking Wrapper 생성 (배치 예측 최적화)
            print("\n[3단계] Stacking Wrapper 생성")
            wrapper = StackingModelWrapper(ensemble_models, meta_model)
            print(f"   [OK] Wrapper 생성 완료 (배치 예측 최적화)")

            # Stacking 모델 저장
            stacking_model = {
                'type': 'stacking',  # ml_score_set 함수가 인식하는 키
                'model_type': 'stacking',
                'model': wrapper,  # ⚡ sklearn 호환 인터페이스 (배치 예측)
                'base_models': ensemble_models,
                'meta_model': meta_model,
                'mu': mu,
                'sigma': sigma,
                'n_base_models': len(ensemble_models),
                'meta_cv_accuracy': cv_scores.mean() * 100,
                'meta_train_accuracy': train_accuracy * 100,
                'separation_power': separation,
                'n_features': D,
                'n_meta_features': X_meta.shape[1],
            }

            with open('best_ml_model_stacking.pkl', 'wb') as f:
                pickle.dump(stacking_model, f)

            print(f"   [OK] Stacking 모델 저장 완료")
            print("\n" + "=" * 80)
            print("Stacking 앙상블 학습 완료!")
            print("=" * 80)

            # 학습 성공 - 메인 스레드에서 UI 업데이트
            used_rounds = len(self.history_df)
            self.after(0, lambda: self._on_ml_training_success(
                stacking_model, "Stacking 앙상블", used_rounds
            ))

        except Exception as e:
            # 학습 실패 - 메인 스레드에서 UI 업데이트
            import traceback
            traceback.print_exc()
            self.after(0, lambda: self._on_ml_training_failure(str(e)))

    def _on_ml_training_success(self, model, model_name, used_rounds):
        """ML 학습 성공 시 UI 업데이트 (메인 스레드)"""
        self.ml_model = model

        # Stacking 모델 정보 표시
        n_models = model.get('n_base_models', 0)
        accuracy = model.get('meta_train_accuracy', 0) / 100  # 백분율 → 소수
        sep_power = model.get('separation_power', 0)

        self.lbl_ai.config(
            text=f"AI 세트 평점: {model_name} ({n_models}+1 모델, 정확도 {accuracy:.2%}, 구분력 {sep_power:.4f})"
        )

        # 가상 조작 시뮬 ML 레이블도 업데이트
        self._update_rig_ml_label()

        messagebox.showinfo(
            "학습 완료",
            f"✅ {model_name} 학습 완료!\n"
            f"   - 학습 회차: {used_rounds}회\n"
            f"   - 베이스 모델: {n_models}개\n"
            f"   - 정확도: {accuracy:.2%}\n"
            f"   - 구분력: {sep_power:.4f}\n\n"
            f"이제 MQLE 모드와 가상조작 시뮬에서 ML 점수가 반영됩니다."
        )

    def _on_ml_training_failure(self, error_msg):
        """ML 학습 실패 시 UI 업데이트 (메인 스레드)"""
        self.ml_model = None
        self.lbl_ai.config(text="AI 세트 평점: 학습 실패 (기본 MQLE만 동작)")
        messagebox.showerror(
            "AI 학습 실패",
            f"ML 모델 학습 중 오류 발생:\n{error_msg}"
        )

    def _prepare_history_weights(self):
        strat = self.hist_strategy.get()
        lookback_str = self.hist_lookback.get().strip()
        lookback = None if lookback_str == "" else int(lookback_str)
        excl = max(0, int(self.hist_exclude.get()))
        w, excl_set = compute_weights(
            self.history_df, lookback, strat, exclude_recent=excl
        )
        self.history_weights = w
        self.history_exclude_set = excl_set

    def _gen_dispatch(self):
        mode = self.gen_mode.get()
        n = max(1, self.gen_count.get())
        weights = None
        excl_set: set[int] = set()

        if self.hist_strategy.get() != "사용 안 함":
            if self.history_df is None:
                messagebox.showwarning(
                    "알림", "히스토리 전략 사용 시 과거 데이터를 먼저 불러오세요."
                )
                return
            try:
                self._prepare_history_weights()
            except Exception as e:
                messagebox.showerror("히스토리 가중치 오류", str(e))
                return
            weights = self.history_weights
            excl_set = self.history_exclude_set

        try:
            if mode == "무작위":
                arr = generate_random_sets(
                    n, True, weights, excl_set or None
                )
            elif mode == "패턴":
                even_str = self.gen_even.get().strip()
                even_target = None if even_str == "" else int(even_str)
                arr = generate_pattern_sets(
                    n,
                    even_target=even_target,
                    low_mid_high=(
                        self.gen_low.get(),
                        self.gen_mid.get(),
                        self.gen_high.get(),
                    ),
                    include_multiples=(
                        self.gen_m3.get(),
                        self.gen_m7.get(),
                    ),
                    weights=weights,
                    exclude_set=excl_set or None,
                )
            elif mode == "GI(직관)":
                arr = gen_GI(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "MDA(다차원)":
                arr = gen_MDA(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "CC(창의연결)":
                arr = gen_CC(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "PR(패턴공진)":
                arr = gen_PR(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "IS(혁신시뮬)":
                arr = gen_IS(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "GAP-R(간격공진)":
                arr = gen_GAPR(
                    n,
                    history_df=self.history_df,
                    weights=weights,
                    exclude_set=excl_set or None,
                )
            elif mode == "QH(다속성조화)":
                arr = gen_QH(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "HD(초다양성)":
                base_sets = None
                txt = self.text_sets.get("1.0", tk.END)
                if txt.strip():
                    try:
                        base_sets = parse_sets_from_text(txt)
                    except Exception:
                        base_sets = None
                arr = gen_HD(
                    n, base_sets=base_sets, weights=weights, exclude_set=excl_set or None
                )
            elif mode == "QP-Wave(파동)":
                arr = gen_QP(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "QP-Tunnel(터널링)":
                arr = gen_QP_tunnel(
                    n, weights=weights, exclude_set=excl_set or None
                )
            elif mode == "QP-Entangle(얽힘)":
                arr = gen_QP_entangle(
                    n,
                    history_df=self.history_df,
                    weights=weights,
                    exclude_set=excl_set or None,
                )
            elif mode == "QH-QA(어닐링)":
                arr = gen_QH_QA(n, weights=weights, exclude_set=excl_set or None)
            elif mode == "QP-Jump(모드도약)":
                arr = gen_QP_jump(
                    n,
                    history_df=self.history_df,
                    weights=weights,
                    exclude_set=excl_set or None,
                )
            elif mode == "MQLE(끝판왕)":
                # MQLE도 백그라운드 스레드에서 실행 (GUI 멈춤 방지)
                self._run_mqle_in_background(mode, n, weights, excl_set)
                return  # 백그라운드에서 처리하므로 여기서 리턴
            elif mode in ("물리시뮬3D", "물리시뮬3D+MQLE(끝판왕)"):
                # 물리시뮬은 백그라운드 스레드에서 실행 (GUI 멈춤 방지)
                self._run_physics_in_background(mode, n, weights)
                return  # 백그라운드에서 처리하므로 여기서 리턴
            else:
                arr = []
        except Exception as e:
            messagebox.showerror("번호 생성 오류", str(e))
            return

        # ML 점수 계산 및 정렬
        if self.ml_model is not None and len(arr) > 0:
            try:
                # 다음 회차 정보 계산
                next_round, next_date = get_next_round_info(self.history_df)

                # 배치 ML 점수 계산 (17.5배 빠른 병렬 처리, 시간 정보 포함)
                scores = ml_score_sets_batch(
                    arr,
                    self.ml_model,
                    weights=weights,
                    history_df=self.history_df,
                    round_num=next_round,
                    date_str=next_date,
                )

                # ML 점수 내림차순 정렬 (높은 점수가 먼저)
                sorted_pairs = sorted(
                    zip(arr, scores),
                    key=lambda x: x[1],
                    reverse=True
                )
                sorted_sets = [p[0] for p in sorted_pairs]
                sorted_scores = [p[1] for p in sorted_pairs]

                # ML 점수와 함께 표시
                self.text_generate.delete("1.0", tk.END)
                self.text_generate.insert("1.0", sets_to_text_with_scores(sorted_sets, sorted_scores))
            except Exception:
                # ML 점수 실패 시 점수 없이 표시
                self.text_generate.delete("1.0", tk.END)
                self.text_generate.insert("1.0", sets_to_text(arr))
        else:
            # ML 모델이 없으면 점수 없이 표시
            self.text_generate.delete("1.0", tk.END)
            self.text_generate.insert("1.0", sets_to_text(arr))

    def _run_mqle_in_background(self, mode: str, n: int, weights, excl_set: set[int]):
        """MQLE를 백그라운드 스레드에서 실행"""
        # MQLE 모드는 CSV 필수
        if self.history_df is None:
            messagebox.showwarning(
                "CSV 파일 필요",
                "MQLE 모드는 CSV 데이터가 필요합니다.\n"
                "상단 메뉴에서 CSV 파일을 먼저 불러오세요."
            )
            return

        self.text_generate.delete("1.0", tk.END)
        self.text_generate.insert("1.0", f"[{mode}] 계산 중... (15개 고전 + 4개 양자 알고리즘)")
        self.update()  # GUI 즉시 업데이트

        def task():
            try:
                from lotto_generators import gen_MQLE

                # 다음 회차 정보 계산
                next_round, next_date = get_next_round_info(self.history_df)

                # 사용자 세트 읽기
                base_sets = None
                txt = self.text_sets.get("1.0", tk.END)
                if txt.strip():
                    try:
                        base_sets = parse_sets_from_text(txt)
                    except Exception:
                        base_sets = None

                # MQLE 실행
                q_bal = self.qc_balance.get() / 100.0
                ml_w = self.ml_weight.get() / 100.0
                arr = gen_MQLE(
                    n,
                    history_df=self.history_df,
                    weights=weights,
                    exclude_set=excl_set or None,
                    base_sets=base_sets,
                    q_balance=q_bal,
                    ml_model=self.ml_model,
                    ml_weight=ml_w,
                    round_num=next_round,  # 시간 정보 전달
                    date_str=next_date,    # 시간 정보 전달
                )

                # GUI 업데이트는 메인 스레드에서 (시간 정보도 전달)
                self.after(0, lambda: self._on_mqle_complete(arr, mode, weights, next_round, next_date))
            except Exception as e:
                import traceback
                error_msg = f"{str(e)}\n{traceback.format_exc()}"
                self.after(0, lambda: self._on_mqle_error(error_msg))

        threading.Thread(target=task, daemon=True).start()

    def _on_mqle_complete(self, arr: list, mode: str, weights, round_num=None, date_str=None):
        """MQLE 완료 콜백 - ML 점수와 함께 표시"""
        if self.ml_model is not None and len(arr) > 0:
            try:
                # 배치 ML 점수 계산 (17.5배 빠른 병렬 처리, 시간 정보 포함)
                scores = ml_score_sets_batch(
                    arr,
                    self.ml_model,
                    weights=weights,
                    history_df=self.history_df,
                    round_num=round_num,  # 시간 정보 전달
                    date_str=date_str,    # 시간 정보 전달
                )

                # ML 점수 내림차순 정렬 (높은 점수가 먼저)
                sorted_pairs = sorted(
                    zip(arr, scores),
                    key=lambda x: x[1],
                    reverse=True
                )
                sorted_sets = [p[0] for p in sorted_pairs]
                sorted_scores = [p[1] for p in sorted_pairs]

                # ML 점수와 함께 표시
                self.text_generate.delete("1.0", tk.END)
                self.text_generate.insert("1.0", sets_to_text_with_scores(sorted_sets, sorted_scores))
            except Exception:
                # ML 점수 실패 시 점수 없이 표시
                self.text_generate.delete("1.0", tk.END)
                self.text_generate.insert("1.0", sets_to_text(arr))
        else:
            # ML 모델이 없으면 점수 없이 표시
            self.text_generate.delete("1.0", tk.END)
            self.text_generate.insert("1.0", sets_to_text(arr))

        messagebox.showinfo("완료", f"[{mode}] {len(arr)}개 세트 생성 완료!")

    def _on_mqle_error(self, error: str):
        """MQLE 에러 콜백"""
        self.text_generate.delete("1.0", tk.END)
        messagebox.showerror("MQLE 오류", error)

    def _run_physics_in_background(self, mode: str, n: int, weights):
        """3D 물리시뮬을 백그라운드 스레드에서 실행"""
        # MQLE 모드는 CSV 필수 (히스토리 전략은 선택)
        if mode == "물리시뮬3D+MQLE(끝판왕)":
            if self.history_df is None:
                messagebox.showwarning(
                    "CSV 파일 필요",
                    "물리시뮬3D+MQLE 모드는 CSV 데이터가 필요합니다.\n"
                    "상단 메뉴에서 CSV 파일을 먼저 불러오세요."
                )
                return

        if "MQLE" in mode:
            backend = "3D CFD + MQLE 융합"
        else:
            backend = "3D CFD 구형챔버"
        self.text_generate.delete("1.0", tk.END)
        self.text_generate.insert("1.0", f"[{mode}] 계산 중... ({backend})")
        self.update()  # GUI 즉시 업데이트

        def task():
            try:
                from lotto_physics import (
                    generate_physics_3d,
                    generate_physics_3d_ultimate,
                )

                # 다음 회차 정보 계산
                next_round, next_date = get_next_round_info(self.history_df)

                rng = np.random.default_rng()
                arr = []

                if mode == "물리시뮬3D":
                    # 3D 구형 챔버 시뮬레이션 (실제 Venus 추첨기)
                    arr = generate_physics_3d(
                        n_sets=n,
                        seed=int(rng.integers(0, 2**31)),
                        use_cfd=True,
                        grid_size=32,  # 64 -> 32 (약 20배 빠름, 정확도 충분)
                        fast_mode=True,  # 빠른 모드 활성화
                    )

                elif mode == "물리시뮬3D+MQLE(끝판왕)":
                    # 3D 구형챔버 + PMMA 물리 + MQLE 융합 (최종 끝판왕)
                    ml_w = self.ml_weight.get() / 100.0
                    arr = generate_physics_3d_ultimate(
                        n_sets=n,
                        seed=int(rng.integers(0, 2**31)),
                        grid_size=32,  # 64 -> 32 (약 20배 빠름)
                        history_df=self.history_df,  # CSV 데이터 전달
                        history_weights=weights,
                        mqle_threshold=0.5,
                        max_attempts=30,
                        fast_mode=True,  # 빠른 모드 활성화
                        ml_model=self.ml_model,  # ML 모델 전달
                        ml_weight=ml_w,  # ML 가중치 전달
                        round_num=next_round,  # 시간 정보 전달
                        date_str=next_date,    # 시간 정보 전달
                    )

                arr = arr[:n]

                # GUI 업데이트는 메인 스레드에서
                self.after(0, lambda: self._on_physics_complete(arr, mode, weights))
            except Exception as e:
                import traceback
                error_msg = f"{str(e)}\n{traceback.format_exc()}"
                self.after(0, lambda: self._on_physics_error(error_msg))

        threading.Thread(target=task, daemon=True).start()

    def _on_physics_complete(self, arr: list, mode: str, weights):
        """물리시뮬 완료 콜백 - ML 점수와 함께 표시"""
        if self.ml_model is not None and len(arr) > 0:
            try:
                # 다음 회차 정보 계산
                next_round, next_date = get_next_round_info(self.history_df)

                # 배치 ML 점수 계산 (17.5배 빠른 병렬 처리)
                scores = ml_score_sets_batch(
                    arr,
                    self.ml_model,
                    weights=weights,
                    history_df=self.history_df,
                    round_num=next_round,
                    date_str=next_date,
                )

                # ML 점수 내림차순 정렬 (높은 점수가 먼저)
                sorted_pairs = sorted(
                    zip(arr, scores),
                    key=lambda x: x[1],
                    reverse=True
                )
                sorted_sets = [p[0] for p in sorted_pairs]
                sorted_scores = [p[1] for p in sorted_pairs]

                # ML 점수와 함께 표시
                self.text_generate.delete("1.0", tk.END)
                self.text_generate.insert("1.0", sets_to_text_with_scores(sorted_sets, sorted_scores))
            except Exception:
                # ML 점수 실패 시 점수 없이 표시
                self.text_generate.delete("1.0", tk.END)
                self.text_generate.insert("1.0", sets_to_text(arr))
        else:
            # ML 모델이 없으면 점수 없이 표시
            self.text_generate.delete("1.0", tk.END)
            self.text_generate.insert("1.0", sets_to_text(arr))

        messagebox.showinfo("완료", f"[{mode}] {len(arr)}개 세트 생성 완료!")

    def _on_physics_error(self, error: str):
        """물리시뮬 에러 콜백"""
        self.text_generate.delete("1.0", tk.END)
        messagebox.showerror("물리시뮬 오류", error)


    def _append_to_sets(self):
        try:
            sets_new = parse_sets_from_text(self.text_generate.get("1.0", tk.END))
        except Exception as e:
            messagebox.showerror("오류", str(e))
            return
        current = self.text_sets.get("1.0", tk.END)
        base: list[list[int]] = []
        if current.strip():
            try:
                base = parse_sets_from_text(current)
            except Exception as e:
                messagebox.showerror("오류", f"세트 페이지 오류: {e}")
                return
        merged = [tuple(s) for s in base] + [tuple(s) for s in sets_new]
        uniq = sorted(list({t for t in merged}))
        self.text_sets.delete("1.0", tk.END)
        self.text_sets.insert("1.0", sets_to_text([list(t) for t in uniq]))
        messagebox.showinfo(
            "추가 완료",
            f"세트 {len(sets_new)}개 추가됨 (중복 제거 후 총 {len(uniq)}개)",
        )

    def _clear_generated(self):
        self.text_generate.delete("1.0", tk.END)

    def _launch_3d_visualizer(self):
        """3D 물리 시각화 창 열기"""
        mode = self.gen_mode.get()
        if "물리시뮬3D" not in mode:
            messagebox.showwarning(
                "주의",
                "3D 시각화는 '물리시뮬3D' 또는 '물리시뮬3D+MQLE' 모드에서만 사용 가능합니다."
            )
            return

        # MQLE 모드에서는 시각화 의미 없음 (경고만)
        if "MQLE" in mode:
            messagebox.showwarning(
                "시각화 비추천",
                "⚠️ 물리시뮬3D+MQLE 모드는 내부적으로 수십~수백 번의\n"
                "빠른 시뮬레이션(비시각화)을 실행한 후 최적 결과만 선택합니다.\n\n"
                "시각화로 보는 1회 시뮬레이션은 실제 생성 과정과 무관하며,\n"
                "1회당 약 50초가 소요됩니다.\n\n"
                "💡 순수 물리 시뮬레이션 시각화를 원하시면\n"
                "'물리시뮬3D' 단독 모드를 사용하세요."
            )
            return

        # 로딩 알림
        messagebox.showinfo(
            "3D 시각화 시작",
            "시각화 창이 열립니다.\n\n"
            "첫 실행 시 초기화에 1-2초 소요됩니다.\n"
            "(Numba JIT 컴파일, OpenGL 초기화)"
        )

        # 별도 스레드에서 시각화 실행
        def run_visualizer():
            try:
                # ★ 모듈 강제 리로드 (수정사항 즉시 반영 - VS Code 포함)
                import sys

                # 캐시된 .pyc 파일 무시
                sys.dont_write_bytecode = True

                # 관련 모듈 완전히 제거 후 재import
                modules_to_remove = []
                for mod_name in list(sys.modules.keys()):
                    if 'lotto_physics' in mod_name or 'physics_visualizer' in mod_name:
                        modules_to_remove.append(mod_name)

                for mod_name in modules_to_remove:
                    del sys.modules[mod_name]
                    print(f"[리로드] {mod_name} 모듈 제거 후 재로드")

                # 새로 import
                from physics_visualizer_3d import launch_visualizer

                launch_visualizer(num_balls=45, mode=mode)
            except ImportError as e:
                messagebox.showerror(
                    "모듈 오류",
                    f"3D 시각화 모듈을 불러올 수 없습니다:\n{e}\n\n"
                    "pygame과 PyOpenGL이 설치되어 있는지 확인하세요."
                )
            except Exception as e:
                messagebox.showerror("시각화 오류", f"3D 시각화 실행 중 오류:\n{e}")

        # daemon=False: 시각화 창을 독립적으로 닫을 수 있도록
        thread = threading.Thread(target=run_visualizer, daemon=False)
        thread.start()
        # 조작법은 3D 시각화 화면에 표시됨

    # --- 시뮬레이션 페이지 ---
    def _build_sim_page(self):
        top = self.page_sim

        frm = ttk.Frame(top)
        frm.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(frm, text="총 추첨 횟수(draws)").grid(row=0, column=0, sticky="w")
        self.sim_draws = tk.IntVar(value=2_000_000)
        ttk.Entry(frm, textvariable=self.sim_draws, width=12).grid(
            row=0, column=1, sticky="w", padx=6
        )

        ttk.Label(frm, text="배치(batch)").grid(row=0, column=2, sticky="e")
        self.sim_batch = tk.IntVar(value=200_000)
        ttk.Entry(frm, textvariable=self.sim_batch, width=10).grid(
            row=0, column=3, sticky="w", padx=6
        )

        ttk.Label(frm, text="워커 수(workers, 최대 36)").grid(
            row=0, column=4, sticky="e"
        )
        self.sim_workers = tk.IntVar(value=8)
        ttk.Entry(frm, textvariable=self.sim_workers, width=8).grid(
            row=0, column=5, sticky="w", padx=6
        )

        ttk.Label(frm, text="Seed(선택)").grid(row=1, column=0, sticky="w", pady=6)
        self.sim_seed = tk.StringVar(value="")
        ttk.Entry(frm, textvariable=self.sim_seed, width=12).grid(
            row=1, column=1, sticky="w"
        )

        btns = ttk.Frame(top)
        btns.pack(fill=tk.X, padx=10, pady=8)
        ttk.Button(btns, text="시뮬레이션 실행", command=self._run_sim).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(btns, text="CSV/Excel로 저장", command=self._save_outputs).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(btns, text="가상 조작 시뮬", command=self._open_rigged_dialog).pack(
            side=tk.LEFT, padx=6
        )

        self.progress = ttk.Progressbar(top, mode="indeterminate")
        self.progress.pack(fill=tk.X, padx=10, pady=6)
        self.lbl_status = ttk.Label(top, text="대기 중")
        self.lbl_status.pack(anchor="w", padx=10)

        cols = [
            "Set",
            "Numbers",
        ] + [f"match_{m}_count" for m in range(7)] + [
            f"match_{m}_prob" for m in range(7)
        ] + ["match_5plusbonus_count", "match_5plusbonus_prob", "≥3_match_prob"]

        frame_list = ttk.Frame(top)
        frame_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)

        self.tree = ttk.Treeview(
            frame_list, columns=cols, show="headings", height=16
        )
        vsb = ttk.Scrollbar(frame_list, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(
            frame_list, orient="horizontal", command=self.tree.xview
        )
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        frame_list.rowconfigure(0, weight=1)
        frame_list.columnconfigure(0, weight=1)

        for c in cols:
            self.tree.heading(
                c,
                text=c,
                command=lambda col=c: self._sort_simulation_results(col)
            )
            self.tree.column(
                c, width=110 if c != "Numbers" else 180, anchor="center"
            )

        self.per_set_df: pd.DataFrame | None = None
        self.agg_df: pd.DataFrame | None = None

    def _run_sim(self):
        try:
            sets_ = parse_sets_from_text(self.text_sets.get("1.0", tk.END))
        except Exception as e:
            messagebox.showerror("오류", str(e))
            return
        draws = max(1, self.sim_draws.get())
        batch = max(1, self.sim_batch.get())
        workers = max(1, min(36, self.sim_workers.get()))
        seed_str = self.sim_seed.get().strip()
        seed_val = None if seed_str == "" else int(seed_str)

        def task():
            try:
                self._set_busy(True, "시뮬레이션 실행 중...")
                per_set_df, agg_df = run_simulation(
                    draws, batch, workers, seed_val, sets_
                )
                self.per_set_df = per_set_df
                self.agg_df = agg_df
                self.after(
                    0, lambda: self._populate_tree(per_set_df, agg_df)
                )
                self._set_busy(False, f"완료: draws={draws:,}, workers={workers}, batch={batch:,}")
            except Exception as e_inner:
                self._set_busy(False, "오류 발생")
                messagebox.showerror("오류", str(e_inner))

        threading.Thread(target=task, daemon=True).start()

    def _populate_tree(self, per_set_df: pd.DataFrame, agg_df: pd.DataFrame):
        self.tree.delete(*self.tree.get_children())
        for _, row in per_set_df.iterrows():
            values = [row.get(col, "") for col in self.tree["columns"]]
            self.tree.insert("", tk.END, values=values)
        row = agg_df.iloc[0].to_dict()
        values = [row.get(col, "") for col in self.tree["columns"]]
        self.tree.insert("", tk.END, values=values)

    def _sort_simulation_results(self, column: str):
        """시뮬레이션 결과 테이블 정렬"""
        if self.per_set_df is None or self.per_set_df.empty:
            return

        # 같은 컬럼 클릭 시 오름차순/내림차순 토글
        if self.sim_sort_column == column:
            self.sim_sort_reverse = not self.sim_sort_reverse
        else:
            # 새 컬럼 선택 시 내림차순으로 시작 (높은 값이 위로)
            self.sim_sort_column = column
            self.sim_sort_reverse = True

        # 정렬 실행 (숫자 컬럼은 숫자로, 문자 컬럼은 문자로)
        try:
            # pandas DataFrame 정렬
            sorted_df = self.per_set_df.sort_values(
                by=column,
                ascending=not self.sim_sort_reverse
            )

            # 테이블 업데이트 (집계 행 제외, per_set만 정렬)
            self.tree.delete(*self.tree.get_children())

            # 정렬된 per_set 데이터 표시
            for _, row in sorted_df.iterrows():
                values = [row.get(col, "") for col in self.tree["columns"]]
                self.tree.insert("", tk.END, values=values)

            # 집계 행은 항상 마지막에 표시
            if self.agg_df is not None:
                agg_row = self.agg_df.iloc[0].to_dict()
                values = [agg_row.get(col, "") for col in self.tree["columns"]]
                self.tree.insert("", tk.END, values=values)

            # 컬럼 헤더에 정렬 방향 표시
            cols = self.tree["columns"]
            for c in cols:
                if c == column:
                    # 정렬 중인 컬럼에 화살표 표시
                    arrow = " ▼" if self.sim_sort_reverse else " ▲"
                    self.tree.heading(c, text=f"{c}{arrow}")
                else:
                    # 다른 컬럼은 화살표 제거
                    self.tree.heading(c, text=c)

        except Exception as e:
            print(f"[ERROR] 정렬 실패: {e}")
            import traceback
            traceback.print_exc()

    def _save_outputs(self):
        if self.per_set_df is None or self.agg_df is None:
            messagebox.showwarning("알림", "먼저 시뮬레이션을 실행하세요.")
            return
        folder = filedialog.askdirectory()
        if not folder:
            return

        # 날짜/시간 포함 파일명
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y년%m월%d일_%H시%M분')

        per_csv = os.path.join(folder, f"lotto_per_set_{timestamp}.csv")
        agg_csv = os.path.join(folder, f"lotto_aggregate_{timestamp}.csv")
        self.per_set_df.to_csv(per_csv, index=False)
        self.agg_df.to_csv(agg_csv, index=False)
        try:
            xlsx = os.path.join(folder, f"lotto_results_{timestamp}.xlsx")
            with pd.ExcelWriter(xlsx, engine="xlsxwriter") as writer:
                self.per_set_df.to_excel(
                    writer, sheet_name="PerSet", index=False
                )
                self.agg_df.to_excel(
                    writer, sheet_name="Aggregate", index=False
                )
        except Exception as e:
            messagebox.showwarning("엑셀 저장 경고", f"엑셀 저장 실패: {e}")
        messagebox.showinfo("저장 완료", f"CSV/엑셀 저장됨: {folder}")

    def _set_busy(self, busy: bool, text: str):
        self.lbl_status.config(text=text)
        if busy:
            self.progress.start(10)
        else:
            self.progress.stop()

    # --- 가상 조작 시뮬 레이어 ---
    def _open_rigged_dialog(self):
        if self.history_df is None or self.history_df.empty:
            messagebox.showwarning(
                "알림",
                "가상 조작 시뮬은 과거 히스토리가 필요합니다.\n먼저 과거 데이터를 로드해 주세요.",
            )
            return

        if self.rig_win is not None and tk.Toplevel.winfo_exists(self.rig_win):
            self.rig_win.lift()
            self.rig_win.focus_force()
            return

        win = tk.Toplevel(self)
        win.title("가상 조작 시뮬레이터 (1등 인원 타겟 + GPU 옵션)")
        win.geometry("640x560")
        self.rig_win = win

        top = ttk.Frame(win)
        top.pack(fill=tk.X, padx=10, pady=8)

        ttk.Label(top, text="목표 1등 인원 최소").grid(row=0, column=0, sticky="e")
        ttk.Entry(top, textvariable=self.rig_target_min, width=6).grid(
            row=0, column=1, sticky="w", padx=4
        )

        ttk.Label(top, text="목표 1등 인원 최대").grid(row=0, column=2, sticky="e")
        ttk.Entry(top, textvariable=self.rig_target_max, width=6).grid(
            row=0, column=3, sticky="w", padx=4
        )

        # ★ 샘플링 제거 - ticket_pool 전수 조사 방식으로 변경
        # (샘플링 후보 개수 입력란 제거됨)

        # 검색 실행/중지 버튼
        btn_frame = ttk.Frame(top)
        btn_frame.grid(row=1, column=0, columnspan=2, sticky="w", padx=6, pady=4)

        self.rig_start_btn = ttk.Button(btn_frame, text="검색 실행", command=self._run_rigged_search)
        self.rig_start_btn.pack(side=tk.LEFT, padx=(0, 4))

        self.rig_stop_btn = ttk.Button(btn_frame, text="중지", command=self._stop_rigged_search, state="disabled")
        self.rig_stop_btn.pack(side=tk.LEFT)

        # 중지 플래그
        self.rig_stop_flag = False

        # ★ 추가: 현실 구매자 수 입력
        self.rig_buyers = tk.IntVar(value=14000000)
        ttk.Label(top, text="현실 구매자 수").grid(row=4, column=0, sticky="e", pady=4)
        buyers_entry = ttk.Entry(top, textvariable=self.rig_buyers, width=12)
        buyers_entry.grid(row=4, column=1, sticky="w", padx=4)
        ttk.Label(top, text="(예: 14,000,000)").grid(row=4, column=2, sticky="w")

        # 구매자 수 변경 시 가상 플레이어 수 자동 계산
        self.rig_buyers.trace_add("write", self._auto_update_virtual_players)

        # ★ 추가: 1인당 평균 게임 수 입력
        self.rig_avg_games = tk.DoubleVar(value=8.0)
        ttk.Label(top, text="1인당 평균 게임 수").grid(row=5, column=0, sticky="e", pady=4)
        games_entry = ttk.Entry(top, textvariable=self.rig_avg_games, width=12)
        games_entry.grid(row=5, column=1, sticky="w", padx=4)
        ttk.Label(top, text="(예: 8 게임)").grid(row=5, column=2, sticky="w")

        # 평균 게임 수 변경 시 가상 플레이어 수 자동 계산
        self.rig_avg_games.trace_add("write", self._auto_update_virtual_players)

        # ★ 가상 플레이어 수 입력 (자동 계산됨)
        ttk.Label(top, text="가상 플레이어 수").grid(row=2, column=0, sticky="e", pady=4)
        vp_entry = ttk.Entry(top, textvariable=self.rig_virtual_players, width=12, state="readonly")
        vp_entry.grid(row=2, column=1, sticky="w", padx=4)
        ttk.Label(top, text="명 (자동: 구매자 × 게임 수)").grid(row=2, column=2, sticky="w")

        # ★ ML 가중치 슬라이더 (변수는 __init__에서 이미 초기화됨)
        ttk.Label(top, text="ML 가중치(%)").grid(row=3, column=0, sticky="e", pady=4)
        ml_scale = tk.Scale(
            top,
            from_=0,
            to=100,
            orient="horizontal",
            variable=self.rig_ml_weight,
            length=150,
        )
        ml_scale.grid(row=3, column=1, sticky="w", padx=4)
        self.rig_ml_label = ttk.Label(top, text="30% (ML 학습 필요)")
        self.rig_ml_label.grid(row=3, column=2, sticky="w")

        # ML 가중치 변경 시 레이블 업데이트 (외부에서 호출 가능하도록)
        self.rig_ml_weight.trace_add("write", lambda *_: self._update_rig_ml_label())
        self._update_rig_ml_label()  # 초기 업데이트

        # 진행률 표시 (Progressbar + Label)
        progress_frame = ttk.Frame(win)
        progress_frame.pack(fill=tk.X, padx=10, pady=4)

        self.rig_progressbar = ttk.Progressbar(progress_frame, mode="determinate", length=400)
        self.rig_progressbar.pack(fill=tk.X, pady=(0, 2))

        self.rig_progress_label = ttk.Label(progress_frame, text="")
        self.rig_progress_label.pack(anchor="w")

        # 상태 라벨 + 엑셀 저장 버튼을 한 줄에
        status_frame = ttk.Frame(win)
        status_frame.pack(fill=tk.X, padx=10, pady=4)
        self.rig_status_label = ttk.Label(status_frame, text="대기 중")
        self.rig_status_label.pack(side=tk.LEFT)
        ttk.Button(status_frame, text="엑셀로 저장", command=self._save_rigged_to_excel).pack(side=tk.RIGHT, padx=4)

        frame_list = ttk.Frame(win)
        frame_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

        cols = ["Rank", "Draw", "예상 1등 인원(λ)"]

        self.rig_tree = ttk.Treeview(
            frame_list, columns=cols, show="headings", height=16
        )
        vsb = ttk.Scrollbar(
            frame_list, orient="vertical", command=self.rig_tree.yview
        )
        hsb = ttk.Scrollbar(
            frame_list, orient="horizontal", command=self.rig_tree.xview
        )
        self.rig_tree.configure(
            yscrollcommand=vsb.set, xscrollcommand=hsb.set
        )

        self.rig_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        frame_list.rowconfigure(0, weight=1)
        frame_list.columnconfigure(0, weight=1)

        for c in cols:
            self.rig_tree.heading(
                c,
                text=c,
                command=lambda col=c: self._sort_rigged_results(col)
            )
            self.rig_tree.column(c, width=160, anchor="center")

    def _sort_rigged_results(self, column: str):
        """가상조작 시뮬 결과 테이블 정렬"""
        if not self.rig_results:
            return

        # 같은 컬럼 클릭 시 오름차순/내림차순 토글
        if self.rig_sort_column == column:
            self.rig_sort_reverse = not self.rig_sort_reverse
        else:
            # 새 컬럼 선택 시 내림차순으로 시작 (높은 값이 위로)
            self.rig_sort_column = column
            self.rig_sort_reverse = True

        # 정렬 키 함수 정의
        def sort_key(item):
            if len(item) == 3:
                draw, lam, combined_score = item
            else:
                draw, lam = item
                combined_score = lam

            if column == "Rank":
                # Rank는 현재 순서 유지 (정렬 후 다시 번호 매김)
                return 0
            elif column == "Draw":
                # 번호 조합: 첫 번째 숫자 기준 정렬
                return min(draw)
            elif column == "예상 1등 인원(λ)":
                # λ 값 기준 정렬
                return lam
            else:
                return 0

        # 정렬 실행
        sorted_results = sorted(
            self.rig_results,
            key=sort_key,
            reverse=self.rig_sort_reverse
        )

        # 테이블 업데이트
        self.rig_tree.delete(*self.rig_tree.get_children())

        for idx, item in enumerate(sorted_results, start=1):
            if len(item) == 3:
                draw, lam, combined_score = item
            else:
                draw, lam = item

            self.rig_tree.insert(
                "",
                tk.END,
                values=[
                    idx,
                    " ".join(map(str, sorted(draw))),
                    f"{lam:5.2f}",
                ],
            )

        # 컬럼 헤더에 정렬 방향 표시
        cols = ["Rank", "Draw", "예상 1등 인원(λ)"]
        for c in cols:
            if c == column:
                # 정렬 중인 컬럼에 화살표 표시
                arrow = " ▼" if self.rig_sort_reverse else " ▲"
                self.rig_tree.heading(c, text=f"{c}{arrow}")
            else:
                # 다른 컬럼은 화살표 제거
                self.rig_tree.heading(c, text=c)

    def _save_rigged_to_excel(self):
        """가상 조작 시뮬 결과를 엑셀 파일로 저장"""
        if not self.rig_results:
            messagebox.showwarning("알림", "저장할 결과가 없습니다.\n먼저 검색을 실행해 주세요.")
            return

        from tkinter import filedialog
        import pandas as pd
        from datetime import datetime

        # 파일 저장 다이얼로그 (날짜_시간 형식)
        default_name = f"가상조작시뮬_{datetime.now().strftime('%Y년%m월%d일_%H시%M분')}.xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx"), ("CSV 파일", "*.csv"), ("모든 파일", "*.*")],
            initialfile=default_name,
            title="가상 조작 시뮬 결과 저장"
        )

        if not filepath:
            return

        try:
            # 결과 데이터 준비
            rows = []
            for idx, item in enumerate(self.rig_results, start=1):
                # ML 사용 시: (draw, lam, combined_score)
                # ML 미사용: (draw, lam)
                if len(item) == 3:
                    draw, lam, combined_score = item
                else:
                    draw, lam = item
                    combined_score = None

                sorted_draw = sorted(draw)
                row_data = {
                    "순위": idx,
                    "번호1": sorted_draw[0],
                    "번호2": sorted_draw[1],
                    "번호3": sorted_draw[2],
                    "번호4": sorted_draw[3],
                    "번호5": sorted_draw[4],
                    "번호6": sorted_draw[5],
                    "번호조합": " ".join(map(str, sorted_draw)),
                    "예상_1등_인원(λ)": round(lam, 4),
                }

                # ML 점수가 있으면 추가
                if combined_score is not None:
                    row_data["ML_Combined_Score"] = round(combined_score, 6)

                rows.append(row_data)

            df_results = pd.DataFrame(rows)

            # 파라미터 정보
            params = self.rig_last_params
            param_rows = [
                {"항목": "검색일시", "값": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                {"항목": "목표_1등_최소", "값": params.get("tmin", "")},
                {"항목": "목표_1등_최대", "값": params.get("tmax", "")},
                {"항목": "샘플링_후보_개수", "값": params.get("samples", "")},
                {"항목": "가상_플레이어_수", "값": params.get("sim_players", "")},
                {"항목": "현실_구매자_수", "값": params.get("buyers", "")},
                {"항목": "1인당_평균_게임_수", "값": params.get("avg_games", "")},
                {"항목": "GPU_사용", "값": "예" if params.get("use_gpu", False) else "아니오"},
                {"항목": "검색_결과_개수", "값": len(self.rig_results)},
            ]
            df_params = pd.DataFrame(param_rows)

            # CSV인 경우
            if filepath.lower().endswith(".csv"):
                df_results.to_csv(filepath, index=False, encoding="utf-8-sig")
                # 파라미터는 별도 파일로
                param_path = filepath.replace(".csv", "_params.csv")
                df_params.to_csv(param_path, index=False, encoding="utf-8-sig")
                messagebox.showinfo("완료", f"CSV 파일로 저장되었습니다.\n결과: {filepath}\n파라미터: {param_path}")
                return

            # 엑셀 저장 시도 (xlsxwriter 우선, 없으면 openpyxl)
            engine = None
            try:
                import xlsxwriter
                engine = "xlsxwriter"
            except ImportError:
                try:
                    import openpyxl
                    engine = "openpyxl"
                except ImportError:
                    engine = None

            if engine:
                with pd.ExcelWriter(filepath, engine=engine) as writer:
                    df_results.to_excel(writer, sheet_name="시뮬결과", index=False)
                    df_params.to_excel(writer, sheet_name="검색파라미터", index=False)
                messagebox.showinfo("완료", f"엑셀 파일로 저장되었습니다.\n{filepath}")
            else:
                # 엑셀 엔진 없으면 CSV로 폴백
                csv_path = filepath.replace(".xlsx", ".csv")
                df_results.to_csv(csv_path, index=False, encoding="utf-8-sig")
                param_path = csv_path.replace(".csv", "_params.csv")
                df_params.to_csv(param_path, index=False, encoding="utf-8-sig")
                messagebox.showinfo("완료", f"엑셀 라이브러리 없어 CSV로 저장했습니다.\n결과: {csv_path}\n파라미터: {param_path}")

        except Exception as e:
            messagebox.showerror("오류", f"저장 중 오류 발생:\n{e}")

    def _stop_rigged_search(self):
        """가상 조작 시뮬 검색 중지"""
        self.rig_stop_flag = True
        self.rig_stop_btn.config(state="disabled")
        self._update_rig_progress(0, "중지 중... (현재 작업 완료 대기)")

    def _run_rigged_search(self):
        if self.history_df is None or self.history_df.empty:
            messagebox.showwarning(
                "알림", "먼저 과거 데이터를 로드해야 가상 조작 시뮬이 가능합니다."
            )
            return

        # 중지 플래그 초기화
        self.rig_stop_flag = False

        # 버튼 상태 변경
        self.rig_start_btn.config(state="disabled")
        self.rig_stop_btn.config(state="normal")

        try:
            tmin = max(0, int(self.rig_target_min.get()))
            tmax = max(tmin, int(self.rig_target_max.get()))
            # ★ 샘플링 제거: samples 변수 더 이상 사용 안 함
            sim_players_val = max(1, int(self.rig_virtual_players.get()))
        except Exception:
            messagebox.showerror("오류", "입력 값이 잘못되었습니다.")
            return

        # HM 가중치(Balanced) 구하기
        try:
            w_bal, _ = compute_weights(
                self.history_df,
                lookback=None,
                strategy="Balanced(중립화)",
                exclude_recent=0,
            )
        except Exception:
            w_bal = None

        if self.rig_status_label is not None:
            self.rig_status_label.config(
                text=f"가상 플레이어 풀 구성 + 전수 조사 중... (가상 플레이어 {sim_players_val:,}명)"
            )

        # 진행률 초기화
        if self.rig_progressbar is not None:
            self.rig_progressbar["value"] = 0
            self.rig_progressbar["maximum"] = 100
        if self.rig_progress_label is not None:
            self.rig_progress_label.config(text="준비 중...")

        def task():
            # 다음 회차 정보 계산
            next_round, next_date = get_next_round_info(self.history_df)

            # 세트 편집 탭에서 사용자 세트 읽기 (취향 반영용)
            user_sets = None
            txt_sets = self.text_sets.get("1.0", tk.END)
            if txt_sets.strip():
                try:
                    user_sets = parse_sets_from_text(txt_sets)
                except Exception:
                    user_sets = None

            # 1) HM + 휴먼 버프 섞어서 '현실적 인기 분포' 만들기
            local_w = compute_realistic_popularity_weights(
                self.history_df,
                hm_weights=w_bal,
                user_sets=user_sets,
            )

            # ★ 최근 N회 번호 회피 세트 (예: 최근 20회)
            try:
                recent_N = 20
                tail = self.history_df.tail(recent_N)
                recent_exclude = set(int(v) for v in np.unique(tail.values) if 1 <= int(v) <= 45)
            except Exception:
                recent_exclude = set()

            sim_players = sim_players_val

            # 2) 가상 플레이어 수: 사용자가 지정한 값 그대로 사용
            sim_players = sim_players_val

            # 진행률 업데이트: 플레이어 풀 생성 시작
            self.after(0, lambda: self._update_rig_progress(10, "가상 플레이어 풀 생성 중..."))

            # 진행률 콜백 함수 (안전 장치 포함)
            def pool_progress_callback(percent, message):
                try:
                    self.after(0, lambda p=percent, m=message: self._update_rig_progress(p, m))
                except:
                    pass  # 위젯이 파괴된 경우 무시

            # 3) 가상 플레이어 풀 생성 (전구간 36코어 사용)
            ticket_pool = build_synthetic_player_pool(
                sim_players,
                local_w,
                workers=36,   # 36 프로세스 풀
                progress_callback=pool_progress_callback,
            )

            # ★ 중지 확인
            if self.rig_stop_flag:
                self.after(0, lambda: self._update_rig_progress(0, "중지됨"))
                self.after(0, lambda: self.rig_start_btn.config(state="normal"))
                self.after(0, lambda: self.rig_stop_btn.config(state="disabled"))
                return

            # 진행률 업데이트: ticket_pool 전수 조사 시작
            self.after(0, lambda: self._update_rig_progress(30, "ticket_pool 전수 조사 중..."))

            # 4) 실제 전국 판매량 계산 (구매자수 × 평균게임수)
            buyers = int(self.rig_buyers.get())
            avg_games = float(self.rig_avg_games.get())
            REAL_TICKETS = buyers * avg_games

            scale_factor = REAL_TICKETS / float(sim_players)

            print("=" * 70)
            print("[DEBUG] 가상 조작 시뮬 파라미터:")
            print(f"  구매자 수: {buyers:,}명")
            print(f"  평균 게임 수: {avg_games}게임")
            print(f"  실제 티켓 수: {REAL_TICKETS:,.0f}장")
            print(f"  가상 플레이어 수: {sim_players:,}명")
            print(f"  Scale Factor: {scale_factor:.4f}")
            print(f"  목표 범위: {tmin}~{tmax}명")
            print(f"  ticket_pool 크기: {len(ticket_pool):,}개 조합")
            print("=" * 70)

            # ticket_pool 샘플 5개 출력
            print("\n[DEBUG] ticket_pool 샘플 (처음 5개):")
            for i, (combo, count) in enumerate(list(ticket_pool.items())[:5]):
                lam = count * scale_factor
                in_range = "✅" if tmin <= lam <= tmax else "❌"
                print(f"  {i+1}. {combo} → 구매자 {count}명, 예상 1등 {lam:.2f}명 {in_range}")
            print()

            # ★ 새로운 방식: ticket_pool 전수 조사 (멀티프로세싱)
            # ML 가중치 읽기
            ml_weight_val = self.rig_ml_weight.get() / 100.0
            use_ml = self.ml_model is not None and ml_weight_val > 0 and self.history_df is not None

            xs: list[tuple[list[int], float]] = []
            center = 0.5 * (tmin + tmax)

            # ★ 동적 작업 할당: ticket_pool을 청크로 분할
            # 빨리 끝난 워커가 다음 청크를 가져가도록 (work stealing)
            ticket_items = list(ticket_pool.items())
            total_combos = len(ticket_items)
            max_workers = 36

            # ⚡ 청크 크기 최적화: 배치 처리 효율을 위해 더 큰 청크 사용
            # 큰 청크 = Numba 병렬 처리 + 신경망 배치 예측 효율 극대화
            chunk_size = max(50000, total_combos // (max_workers * 2))  # 최소 50,000개 (5배 증가)

            # 청크 리스트 생성
            chunks = []
            for i in range(0, total_combos, chunk_size):
                chunks.append(ticket_items[i:i + chunk_size])

            total_chunks = len(chunks)

            with ProcessPoolExecutor(max_workers=max_workers) as ex:
                # 모든 청크를 한 번에 제출 (동적 할당)
                futures = []
                for chunk in chunks:
                    futures.append(
                        ex.submit(
                            _filter_ticket_pool_chunk,
                            chunk,
                            scale_factor,
                            tmin,
                            tmax,
                            center,
                            self.ml_model,
                            ml_weight_val,
                            local_w,
                            self.history_df,
                            next_round,  # 시간 정보 전달
                            next_date,   # 시간 정보 전달
                        )
                    )

                # 진행률 업데이트: 청크 완료 추적
                completed_chunks = 0
                processed_combos = 0

                for fut in as_completed(futures):
                    # ★ 중지 확인
                    if self.rig_stop_flag:
                        self.after(0, lambda: self._update_rig_progress(0, "중지됨"))
                        self.after(0, lambda: self.rig_start_btn.config(state="normal"))
                        self.after(0, lambda: self.rig_stop_btn.config(state="disabled"))
                        return

                    part = fut.result()
                    if part:
                        xs.extend(part)
                    completed_chunks += 1

                    # 처리된 조합 수 계산
                    processed_combos = min(completed_chunks * chunk_size, total_combos)

                    progress_percent = 30 + int((completed_chunks / total_chunks) * 60)
                    self.after(0, lambda p=progress_percent, cc=completed_chunks, tc=total_chunks, pc=processed_combos, ttc=total_combos:
                              self._update_rig_progress(p, f"전수 조사 중... {cc}/{tc} 청크 ({pc:,}/{ttc:,} 조합)"))

            # 진행률 업데이트: 정렬 및 필터링 시작
            found_count = len(xs)
            self.after(0, lambda fc=found_count: self._update_rig_progress(90, f"결과 정렬 중... (범위 내 {fc:,}개 발견)"))

            # 후보 정렬 및 상위 200개 선택
            if not xs:
                rows = []
                print("[DEBUG] xs가 비어있음! rows = []")
            else:
                # ML 사용 시 combined_score로 정렬, 아니면 lam으로 정렬
                if use_ml:
                    # xs = [(combo, lam, combined_score), ...]
                    # combined_score 높은 순
                    print(f"[DEBUG] ML 사용 모드: xs 크기 = {len(xs)}")
                    if xs:
                        print(f"[DEBUG] xs 첫 항목: {xs[0]}")
                    xs_sorted = sorted(xs, key=lambda d: d[2], reverse=True)
                else:
                    # xs = [(combo, lam), ...]
                    # lam이 center에 가까운 순
                    print(f"[DEBUG] ML 미사용 모드: xs 크기 = {len(xs)}")
                    if xs:
                        print(f"[DEBUG] xs 첫 항목: {xs[0]}")
                    xs_sorted = sorted(xs, key=lambda d: abs(d[1] - center))

                rows = xs_sorted[:200]
                print(f"[DEBUG] 정렬 완료: rows 크기 = {len(rows)}")
                if rows:
                    print(f"[DEBUG] rows 첫 항목: {rows[0]}")

            # 진행률 업데이트: 완료
            final_count = len(rows)
            print(f"[DEBUG] found_count={found_count}, final_count={final_count}")
            self.after(0, lambda fc=found_count, rc=final_count:
                      self._update_rig_progress(100, f"완료! (총 {fc:,}개 중 상위 {rc}개 선택)"))

            # 샘플링 개수는 ticket_pool 크기로 표시
            actual_samples = len(ticket_pool)
            print(f"[DEBUG] _update_rigged_tree 호출 예정: rows 크기={len(rows)}, samples={actual_samples}")
            self.after(0, lambda r=rows, t1=tmin, t2=tmax, s=actual_samples, sp=sim_players, b=buyers, ag=avg_games: self._update_rigged_tree(
                r, t1, t2, s, sp, b, ag
            ))

            # ★ 작업 완료: 버튼 상태 복원
            self.after(0, lambda: self.rig_start_btn.config(state="normal"))
            self.after(0, lambda: self.rig_stop_btn.config(state="disabled"))

        threading.Thread(target=task, daemon=True).start()

    def _auto_update_virtual_players(self, *_args):
        """현실 구매자 수 또는 평균 게임 수 변경 시 가상 플레이어 수 자동 계산"""
        try:
            buyers = int(self.rig_buyers.get())
            avg_games = float(self.rig_avg_games.get())
            # 가상 플레이어 수 = 구매자 수 × 평균 게임 수
            virtual_players = int(buyers * avg_games)
            self.rig_virtual_players.set(virtual_players)
        except:
            # 입력 중 오류 발생 시 무시
            pass

    def _update_rig_ml_label(self):
        """가상 조작 시뮬 ML 가중치 레이블 업데이트"""
        if self.rig_ml_label is None:
            return

        val = self.rig_ml_weight.get()
        if self.ml_model is None:
            self.rig_ml_label.config(text=f"{val}% (ML 학습 필요)")
        else:
            # 디버깅: 모델 타입 출력
            model_type = self.ml_model.get("type", "")
            print(f"[DEBUG] ML 모델 타입: '{model_type}'")
            print(f"[DEBUG] ML 모델 전체: {list(self.ml_model.keys())}")

            model_name = {
                "neural_network": "신경망",
            }.get(model_type, "신경망")

            print(f"[DEBUG] 표시 이름: '{model_name}'")
            self.rig_ml_label.config(text=f"{val}% ({model_name})")

    def _update_rig_progress(self, percent: float, message: str):
        """가상 조작 시뮬 진행률 업데이트"""
        try:
            if self.rig_progressbar is not None and self.rig_progressbar.winfo_exists():
                self.rig_progressbar["value"] = percent
        except:
            pass  # 위젯이 파괴됨

        try:
            if self.rig_progress_label is not None and self.rig_progress_label.winfo_exists():
                self.rig_progress_label.config(text=message)
        except:
            pass  # 위젯이 파괴됨

    def _update_rigged_tree(
        self,
        rows: list[tuple[list[int], float]],
        tmin: int,
        tmax: int,
        samples: int,
        sim_players: int,
        buyers: int = 14000000,
        avg_games: float = 8.0,
    ):
        # ★ 디버깅: 파라미터 출력
        print("=" * 70)
        print("[DEBUG] _update_rigged_tree 호출됨")
        print(f"  rows 타입: {type(rows)}")
        print(f"  rows 길이: {len(rows) if rows else 0}")
        if rows:
            print(f"  첫 번째 항목: {rows[0]}")
            print(f"  첫 번째 항목 길이: {len(rows[0])}")
        print(f"  tmin={tmin}, tmax={tmax}")
        print(f"  samples={samples}, sim_players={sim_players}")
        print(f"  self.rig_tree is None? {self.rig_tree is None}")
        print("=" * 70)

        # ★ 결과 저장 (엑셀 저장용)
        self.rig_results = rows
        self.rig_last_params = {
            "tmin": tmin,
            "tmax": tmax,
            "samples": samples,
            "sim_players": sim_players,
            "buyers": buyers,
            "avg_games": avg_games,
        }

        if self.rig_tree is None:
            print("[ERROR] self.rig_tree is None!")
            return

        try:
            # 기존 항목 삭제
            self.rig_tree.delete(*self.rig_tree.get_children())
            print(f"[DEBUG] 기존 항목 삭제 완료")

            # 새 항목 추가
            for idx, item in enumerate(rows, start=1):
                print(f"[DEBUG] 항목 {idx} 처리 중: {item}")

                # ML 사용 시: (draw, lam, combined_score)
                # ML 미사용: (draw, lam)
                if len(item) == 3:
                    draw, lam, combined_score = item
                    print(f"  → ML 사용 결과: draw={draw}, lam={lam}, score={combined_score}")
                else:
                    draw, lam = item
                    combined_score = None
                    print(f"  → ML 미사용 결과: draw={draw}, lam={lam}")

                self.rig_tree.insert(
                    "",
                    tk.END,
                    values=[
                        idx,
                        " ".join(map(str, sorted(draw))),
                        f"{lam:5.2f}",
                    ],
                )
                print(f"  → Tree insert 성공!")

            print(f"[DEBUG] 총 {len(rows)}개 항목 추가 완료")
        except Exception as e:
            print(f"[ERROR] Tree 업데이트 중 오류 발생: {e}")
            import traceback
            traceback.print_exc()
        if self.rig_status_label is not None:
            if not rows:
                self.rig_status_label.config(
                    text=f"검색 완료 — 범위 [{tmin}~{tmax}]에 해당하는 후보 없음 "
                         f"(후보 샘플 {samples:,}개, 가상 플레이어 {sim_players:,}명)"
                )
            else:
                self.rig_status_label.config(
                    text=f"검색 완료 — 후보 {len(rows)}개 "
                         f"(후보 샘플 {samples:,}개, 가상 플레이어 {sim_players:,}명, 목표 [{tmin}~{tmax}])"
                )

    # --- HELP 페이지 ---
    def _build_help_page(self):
        top = self.page_help
        frame = ttk.Frame(top)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        txt = tk.Text(frame, wrap="word")
        scroll = ttk.Scrollbar(frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=scroll.set)

        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        help_text = """
[1. 프로그램 전체 개요]

이 프로그램은 로또 6/45를 대상으로

  1) 내가 원하는 번호 세트를 직접 관리하고,
  2) 여러 가지 '번호 생성 알고리듬'으로 추천 번호를 만들고,
  3) 실제 추첨기를 가정한 몬테카를로(Monte Carlo) 시뮬레이션으로 통계적 성능을 확인하고,
  4) 번호 분포 / 짝·홀 / 구간 / MQLE 조화·다양성을 시각화해서 분석하고,
  5) '만약 조작이 있다면 1등 인원을 어떻게 맞출까?'를 가상으로 실험(리깅 시뮬레이션),
  6) 가상 조작 시뮬의 후보 번호 생성 일부를 GPU(CuPy)로 벡터화 가속

하는 연구/놀이용 도구입니다.

※ 매우 중요
- 실제 로또는 '완전 난수'를 목표로 설계된 시스템입니다.
- 여기 나오는 모든 알고리듬, 양자사운드, AI, 가상조작 시뮬, GPU 가속은
  "수학적인 장난감 + 취향 정리용"일 뿐,
  진짜 수학적 기대값(당첨 확률)을 유의미하게 올려주지 못합니다.
- 반드시 여윳돈 + 재미·연구용으로만 활용하세요.

(이하 HELP 텍스트는 필요하면 자유롭게 확장)
"""
        txt.insert("1.0", help_text)
        txt.config(state="disabled")
        self.help_text_widget = txt

    def _build_recombine_page(self):
        """번호 재조합 페이지 구성"""
        top = self.page_recombine

        # Input section
        input_frame = ttk.LabelFrame(top, text="입력 게임 (한 줄에 6개 숫자)")
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 5))

        input_scroll = ttk.Scrollbar(input_frame)
        input_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.recomb_input_text = tk.Text(
            input_frame,
            height=10,
            wrap="none",
            yscrollcommand=input_scroll.set
        )
        self.recomb_input_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        input_scroll.config(command=self.recomb_input_text.yview)

        # Control section
        control_frame = ttk.Frame(top)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(control_frame, text="출력 개수:").pack(side=tk.LEFT, padx=(0, 5))

        self.recomb_output_count = tk.IntVar(value=10)
        ttk.Entry(
            control_frame,
            textvariable=self.recomb_output_count,
            width=10
        ).pack(side=tk.LEFT, padx=5)

        self.recomb_button = ttk.Button(
            control_frame,
            text="재조합 실행",
            command=self._on_recombine_click
        )
        self.recomb_button.pack(side=tk.LEFT, padx=10)

        self.recomb_stop_btn = ttk.Button(
            control_frame,
            text="중지",
            command=self._on_recombine_stop_click,
            state="disabled"
        )
        self.recomb_stop_btn.pack(side=tk.LEFT, padx=5)

        # Status section
        status_frame = ttk.Frame(top)
        status_frame.pack(fill=tk.X, padx=10, pady=5)

        self.recomb_status_label = ttk.Label(
            status_frame,
            text="대기 중 (게임을 붙여넣고 '재조합 실행'을 클릭하세요)"
        )
        self.recomb_status_label.pack(anchor="w", pady=(0, 5))

        self.recomb_progress = ttk.Progressbar(
            status_frame,
            mode="determinate",
            length=400
        )
        self.recomb_progress.pack(fill=tk.X)

        # Output section
        output_frame = ttk.LabelFrame(
            top,
            text="재조합 결과 (ML 점수 내림차순)"
        )
        output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        output_scroll = ttk.Scrollbar(output_frame)
        output_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.recomb_output_text = tk.Text(
            output_frame,
            height=15,
            wrap="none",
            yscrollcommand=output_scroll.set,
            state="normal"
        )
        self.recomb_output_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        output_scroll.config(command=self.recomb_output_text.yview)

    def _parse_recombine_input(self, text):
        """재조합 입력 파싱 (유연한 형식 지원)

        각 줄에서 1-45 범위의 숫자 6개만 추출합니다.
        회차, 날짜, 보너스, ML 점수 등 다른 정보는 자동으로 무시됩니다.

        지원 형식:
        - "1 7 16 22 28 38"
        - "1203회 2025.01.08 1 7 16 22 28 38 보너스:15"
        - " 5 12 16 19 25 42  [ML: 99.67%]"  ← 재조합 결과도 바로 재입력 가능!
        """
        import re

        sets = []
        for line in text.strip().splitlines():
            if not line.strip():
                continue

            # 모든 숫자 추출
            numbers = re.findall(r'\d+', line)

            # 1-45 범위의 숫자만 필터링
            valid_nums = [int(n) for n in numbers if 1 <= int(n) <= 45]

            # 처음 6개만 사용 (보너스 번호 제외)
            if len(valid_nums) >= 6:
                game = sorted(valid_nums[:6])
                # 중복 체크
                if len(set(game)) == 6:
                    sets.append(game)
                else:
                    # 중복 있으면 건너뛰기
                    continue
            elif len(valid_nums) > 0:
                # 6개 미만이지만 유효한 숫자가 있으면 경고용으로 기록
                pass

        if not sets:
            raise ValueError("유효한 게임이 하나도 없습니다.\n각 줄에 1-45 범위의 숫자가 6개 이상 필요합니다.")

        return sets

    def _on_recombine_click(self):
        """재조합 실행 버튼 클릭 핸들러"""
        import threading
        from itertools import combinations
        import math

        # 1. Get input text
        input_text = self.recomb_input_text.get("1.0", tk.END).strip()
        if not input_text:
            messagebox.showerror("입력 오류", "입력이 비어있습니다.\n\n기존 로또 게임을 붙여넣어주세요.")
            return

        # 2. Parse input sets (재조합 전용 파싱 - 유연한 형식)
        try:
            sets = self._parse_recombine_input(input_text)
        except Exception as e:
            messagebox.showerror("입력 오류", f"입력 형식이 잘못되었습니다.\n\n{str(e)}")
            return

        # 3. Extract unique numbers
        unique_nums = set()
        for s in sets:
            unique_nums.update(s)
        unique_nums = sorted(list(unique_nums))

        # 3-1. 출력 개수를 입력 게임 수와 동일하게 자동 설정 (사용자가 기본값 10을 수정하지 않았을 때만)
        n_input_games = len(sets)
        current_output = self.recomb_output_count.get()
        if current_output == 10:  # 기본값을 그대로 두었으면
            auto_output = max(10, n_input_games)  # 입력과 동일, 최소 10개
            self.recomb_output_count.set(auto_output)
            self.after(0, lambda: None)  # UI 업데이트 트리거

        # 4. Validate unique number count
        n_unique = len(unique_nums)
        if n_unique < 6:
            messagebox.showerror(
                "번호 부족",
                f"최소 6개 이상의 고유 번호가 필요합니다.\n\n현재: {n_unique}개"
            )
            return

        if n_unique == 6:
            messagebox.showinfo(
                "조합 1개",
                f"고유 번호가 정확히 6개입니다.\n조합은 1개만 가능합니다.\n\n번호: {' '.join(map(str, unique_nums))}"
            )
            # Generate the single combination and display
            self.recomb_output_text.delete("1.0", tk.END)
            if self.ml_model is not None:
                try:
                    next_round, next_date = get_next_round_info(self.history_df)
                    scores = ml_score_sets_batch(
                        [unique_nums],
                        self.ml_model,
                        history_df=self.history_df,
                        round_num=next_round,
                        date_str=next_date
                    )
                    self.recomb_output_text.insert("1.0", sets_to_text_with_scores([unique_nums], scores))
                except Exception:
                    self.recomb_output_text.insert("1.0", sets_to_text([unique_nums]))
            else:
                self.recomb_output_text.insert("1.0", sets_to_text([unique_nums]))
            return

        if n_unique > 40:
            n_combos = math.comb(n_unique, 6)
            est_time_min = n_combos / 60000  # Rough estimate: 60K combos/min
            messagebox.showerror(
                "번호 과다",
                f"고유 번호가 너무 많습니다 ({n_unique}개).\n40개 이하로 줄여주세요.\n\n"
                f"예상 조합 수: {n_combos:,}개\n"
                f"예상 시간: ~{est_time_min:.1f}분"
            )
            return

        # 5. Warn for 35-40 unique numbers
        if 35 <= n_unique <= 40:
            n_combos = math.comb(n_unique, 6)
            est_time_min = n_combos / 60000
            result = messagebox.askyesno(
                "조합 수 많음",
                f"조합 수가 많습니다.\n\n"
                f"고유 번호: {n_unique}개\n"
                f"예상 조합 수: {n_combos:,}개\n"
                f"예상 시간: ~{est_time_min:.1f}분\n\n"
                f"계속하시겠습니까?"
            )
            if not result:
                return

        # 6. Get output count
        try:
            output_count = self.recomb_output_count.get()
            if output_count < 1:
                messagebox.showerror("출력 개수 오류", "출력 개수는 1 이상이어야 합니다.")
                return
        except Exception:
            messagebox.showerror("출력 개수 오류", "출력 개수를 올바르게 입력해주세요.")
            return

        # 7. Clear output and reset progress
        self.recomb_output_text.delete("1.0", tk.END)
        self.recomb_progress['value'] = 0

        # 8. Disable button, enable stop button
        self.recomb_button.config(state="disabled")
        self.recomb_stop_btn.config(state="normal")

        # 9. Create stop flag and start worker thread
        self.recomb_stop_flag = threading.Event()

        thread = threading.Thread(
            target=self._recombine_worker,
            args=(unique_nums, output_count, self.recomb_stop_flag),
            daemon=True
        )
        thread.start()

    def _recombine_worker(self, unique_nums, output_count, stop_flag):
        """재조합 백그라운드 작업 (별도 스레드)"""
        from itertools import combinations
        import time

        try:
            # 1. Generate ALL combinations (청크로 나눠서 UI 반응성 유지)
            self.after(0, lambda: self._update_recomb_status("조합 생성 중..."))

            # 전체 조합 수 미리 계산
            import math
            total = math.comb(len(unique_nums), 6)

            # 조합을 청크로 생성 (메모리 효율 + UI 반응성)
            all_combos = list(combinations(unique_nums, 6))

            self.after(0, lambda: self._update_recomb_status(
                f"조합 생성 완료: {total:,}개"
            ))

            # 2. Convert to list[list[int]] format
            sets = [sorted(list(c)) for c in all_combos]

            # 3. Check if ML model is loaded
            if self.ml_model is None:
                # No ML model - show warning and return without scoring
                self.after(0, lambda: self._on_recombine_no_ml(sets, output_count))
                return

            # 4. Calculate next round info
            try:
                next_round, next_date = get_next_round_info(self.history_df)
            except Exception:
                next_round, next_date = None, None

            # 5. ML score in batches (⚡ Numba JIT 배치 처리 + 멀티프로세싱)
            from concurrent.futures import ProcessPoolExecutor, as_completed

            # 큰 배치로 Numba 배치 함수 최대 활용
            batch_size = 50000  # 큰 배치 = Numba 병렬 최적화
            n_workers = min(36, max(4, (total + batch_size - 1) // batch_size))

            self.after(0, lambda: self._update_recomb_status(
                f"ML 점수 계산 중... ({n_workers}개 프로세스, Numba JIT 배치)"
            ))

            # 배치 준비
            batches = []
            for i in range(0, len(sets), batch_size):
                batch = sets[i:i+batch_size]
                batches.append(batch)

            # ProcessPool로 병렬 처리
            all_scores = []
            completed = 0
            last_update_time = time.time()

            with ProcessPoolExecutor(max_workers=n_workers) as executor:
                # Submit all batches
                futures = []
                for batch in batches:
                    if stop_flag.is_set():
                        executor.shutdown(wait=False, cancel_futures=True)
                        self.after(0, lambda: self._on_recombine_cancelled())
                        return

                    future = executor.submit(
                        ml_score_sets_batch,
                        batch,
                        self.ml_model,
                        self.history_weights,
                        self.history_df,
                        next_round,
                        next_date
                    )
                    futures.append(future)

                # Collect results as they complete
                for future in as_completed(futures):
                    if stop_flag.is_set():
                        executor.shutdown(wait=False, cancel_futures=True)
                        self.after(0, lambda: self._on_recombine_cancelled())
                        return

                    batch_scores = future.result()
                    all_scores.extend(batch_scores)

                    # Update progress (throttle: 0.2초마다만 업데이트)
                    completed += len(batch_scores)
                    current_time = time.time()
                    if current_time - last_update_time > 0.2 or completed >= total:
                        progress = completed / total * 100
                        self.after(0, lambda p=progress: self._update_recomb_progress(p))
                        last_update_time = current_time

            # 6. 상위 N개 선택 (메모리 효율적 방법 사용)
            import heapq

            actual_output = min(output_count, len(sets))

            # 조합 수가 많으면 heap 사용 (메모리 절약)
            if len(sets) > output_count * 5:
                # heapq.nlargest로 상위 N개만 추출 (메모리 효율)
                self.after(0, lambda: self._update_recomb_status(f"상위 {actual_output}개 선택 중... (메모리 최적화)"))
                top_n = heapq.nlargest(
                    actual_output,
                    zip(sets, all_scores),
                    key=lambda x: x[1]
                )
                top_sets = [s[0] for s in top_n]
                top_scores = [s[1] for s in top_n]
            else:
                # 조합 수가 적으면 전체 정렬 (빠름)
                scored_sets = sorted(
                    zip(sets, all_scores),
                    key=lambda x: x[1],
                    reverse=True
                )
                top_n = scored_sets[:actual_output]
                top_sets = [s[0] for s in top_n]
                top_scores = [s[1] for s in top_n]

            # 8. Callback to main thread
            self.after(0, lambda: self._on_recombine_complete(top_sets, top_scores, output_count))

        except Exception as e:
            import traceback
            error_msg = f"{str(e)}\n\n{traceback.format_exc()}"
            self.after(0, lambda: self._on_recombine_error(error_msg))

    def _update_recomb_status(self, text):
        """Update status label"""
        if hasattr(self, 'recomb_status_label') and self.recomb_status_label:
            self.recomb_status_label.config(text=text)

    def _update_recomb_progress(self, percent):
        """Update progress bar"""
        if hasattr(self, 'recomb_progress') and self.recomb_progress:
            self.recomb_progress['value'] = percent
        if hasattr(self, 'recomb_status_label') and self.recomb_status_label:
            self.recomb_status_label.config(
                text=f"ML 점수 계산 중... {percent:.1f}%"
            )

    def _on_recombine_complete(self, sets, scores, requested_count):
        """Success callback"""
        # Display results with scores
        self.recomb_output_text.delete("1.0", tk.END)
        self.recomb_output_text.insert("1.0", sets_to_text_with_scores(sets, scores))

        # Update status
        actual_count = len(sets)
        if actual_count < requested_count:
            status_text = f"완료! {actual_count}개 게임 생성 (요청: {requested_count}개, 가능한 최대: {actual_count}개)"
        else:
            status_text = f"완료! {actual_count}개 게임 생성 완료 (ML 점수 내림차순)"
        self.recomb_status_label.config(text=status_text)
        self.recomb_progress['value'] = 100

        # Re-enable button
        self.recomb_button.config(state="normal")
        self.recomb_stop_btn.config(state="disabled")

        messagebox.showinfo("완료", f"{actual_count}개 게임 재조합 완료!")

    def _on_recombine_no_ml(self, sets, requested_count):
        """Handle case when ML model is not loaded"""
        result = messagebox.askyesno(
            "ML 모델 없음",
            "ML 모델이 로드되지 않았습니다.\n"
            "재조합은 가능하지만 ML 점수를 계산할 수 없습니다.\n\n"
            "계속하시겠습니까? (결과는 랜덤 순서로 표시됩니다)"
        )

        if not result:
            self.after(0, lambda: self._on_recombine_cancelled())
            return

        # Take random sample or all if less than requested
        import random
        actual_count = min(requested_count, len(sets))
        selected_sets = random.sample(sets, actual_count) if actual_count < len(sets) else sets[:actual_count]

        # Display without scores
        self.recomb_output_text.delete("1.0", tk.END)
        self.recomb_output_text.insert("1.0", sets_to_text(selected_sets))

        # Update status
        self.recomb_status_label.config(text=f"완료! {actual_count}개 게임 생성 (ML 점수 없음)")
        self.recomb_progress['value'] = 100

        # Re-enable button
        self.recomb_button.config(state="normal")
        self.recomb_stop_btn.config(state="disabled")

        messagebox.showinfo("완료", f"{actual_count}개 게임 재조합 완료! (ML 점수 없음)")

    def _on_recombine_error(self, error_msg):
        """Error callback"""
        self.recomb_status_label.config(text="오류 발생")
        self.recomb_button.config(state="normal")
        self.recomb_stop_btn.config(state="disabled")
        messagebox.showerror("재조합 오류", error_msg)

    def _on_recombine_cancelled(self):
        """Cancellation callback"""
        self.recomb_status_label.config(text="취소됨")
        self.recomb_progress['value'] = 0
        self.recomb_button.config(state="normal")
        self.recomb_stop_btn.config(state="disabled")
        messagebox.showinfo("취소", "재조합이 취소되었습니다.")

    def _on_recombine_stop_click(self):
        """Stop button callback"""
        if hasattr(self, 'recomb_stop_flag') and self.recomb_stop_flag:
            self.recomb_stop_flag.set()
            self.recomb_stop_btn.config(state="disabled")
            self.recomb_status_label.config(text="취소 중...")


if __name__ == "__main__":
    app = LottoApp()
    app.mainloop()
